import csv
import json
import os
import subprocess
import sys
import tempfile
import tkinter as tk
from datetime import datetime, timedelta
from tkinter import filedialog, messagebox, ttk, colorchooser
from tkcalendar import DateEntry

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    Workbook = None
    load_workbook = None
    Alignment = None
    Font = None
    PatternFill = None

# Sprawdzenie ścieżki uruchomienia aplikacji
if getattr(sys, "frozen", False):
    APP_DIR = os.path.dirname(sys.executable)
else:
    APP_DIR = os.path.dirname(os.path.abspath(__file__))

CONFIG_FILE = os.path.join(APP_DIR, "config.json")
DEFAULT_DB_FILE = os.path.join(APP_DIR, "cnc_schedule_db.json")


def get_saved_db_path():
    """Pobiera ścieżkę do bazy z config.json lub zwraca domyślną."""
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                return data.get("db_path", DEFAULT_DB_FILE)
        except Exception:
            pass
    return DEFAULT_DB_FILE


def save_saved_db_path(path):
    """Zapisuje ścieżkę bazy do config.json."""
    try:
        data = {}
        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                    data = json.load(f)
            except Exception:
                pass
        data["db_path"] = path
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
    except Exception as e:
        print(f"Błąd zapisu config.json: {e}")


class ListManagerDialog(tk.Toplevel):
    """Uniwersalne okno dialogowe do zarządzania listami (Klienci, Maszyny, Operatorzy, Statusy)"""
    def __init__(self, parent, title, items_set, callback_on_update):
        super().__init__(parent)
        self.title(title)
        self.geometry("350x400")
        self.resizable(False, False)
        self.grab_set()

        self.items_set = items_set
        self.callback_on_update = callback_on_update

        tk.Label(self, text=title, font=("Arial", 11, "bold")).pack(pady=10)

        # Ramka listy i paska przewijania
        frame_list = tk.Frame(self)
        frame_list.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        self.listbox = tk.Listbox(frame_list, font=("Arial", 10), selectmode=tk.SINGLE)
        scrollbar = tk.Scrollbar(frame_list, orient=tk.VERTICAL, command=self.listbox.yview)
        self.listbox.config(yscrollcommand=scrollbar.set)

        self.listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.refresh_listbox()

        # Ramka wprowadzania nowej pozycji
        frame_add = tk.Frame(self)
        frame_add.pack(fill=tk.X, padx=10, pady=10)

        self.entry_new = tk.Entry(frame_add, font=("Arial", 10))
        self.entry_new.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        self.entry_new.bind("<Return>", lambda e: self.add_item())

        btn_add = tk.Button(frame_add, text="Add", bg="#d4edda", command=self.add_item, width=8)
        btn_add.pack(side=tk.RIGHT)

        # Ramka przycisków akcji
        frame_actions = tk.Frame(self)
        frame_actions.pack(fill=tk.X, padx=10, pady=(0, 10))

        btn_delete = tk.Button(frame_actions, text="Delete Selected", bg="#f8d7da", command=self.delete_item)
        btn_delete.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))

        btn_close = tk.Button(frame_actions, text="Close", command=self.destroy, width=10)
        btn_close.pack(side=tk.RIGHT)

    def refresh_listbox(self):
        self.listbox.delete(0, tk.END)
        for item in sorted(self.items_set):
            self.listbox.insert(tk.END, item)

    def add_item(self):
        val = self.entry_new.get().strip()
        if val:
            if val not in self.items_set:
                self.items_set.add(val)
                self.refresh_listbox()
                self.entry_new.delete(0, tk.END)
                self.callback_on_update()
            else:
                messagebox.showwarning("Warning", "Item already exists!", parent=self)

    def delete_item(self):
        selected_indices = self.listbox.curselection()
        if not selected_indices:
            messagebox.showwarning("Warning", "Please select an item to delete.", parent=self)
            return
        val = self.listbox.get(selected_indices[0])
        if messagebox.askyesno("Confirmation", f"Are you sure you want to delete '{val}'?", parent=self):
            self.items_set.discard(val)
            self.refresh_listbox()
            self.callback_on_update()


class ScheduleApp:
    def __init__(self, root):
        self.root = root
        self.root.title(f"CNC Production Schedule v4.2")
        self.root.geometry("1400x800")

        # Bazy danych (zbiory)
        self.clients_db = set()
        self.machines_db = set()
        self.operators_db = set()
        self.status_db = set()
        self.status_colors = {}
        self.mct_folder = ""
        self.material_register_file = ""
        self.material_register_data = []

        self.db_file = get_saved_db_path()
        self.load_data()

        # Główny układ GUI
        self.create_widgets()
        self.refresh_tree()

    def load_material_register(self):
        """Wczytuje plik material_register.json wskazany w ustawieniach."""
        self.material_register_data = []
        if not self.material_register_file or not os.path.exists(self.material_register_file):
            return

        try:
            with open(self.material_register_file, "r", encoding="utf-8") as f:
                content = json.load(f)
                if isinstance(content, list):
                    self.material_register_data = content
                elif isinstance(content, dict):
                    # Jeśli dane są zagnieżdżone, spróbuj je wyciągnąć
                    for val in content.values():
                        if isinstance(val, list):
                            self.material_register_data = val
                            break
        except Exception as e:
            print(f"Błąd wczytywania rejestru materiałowego: {e}")

    def get_cast_no_for_mct(self, mct_value):
        """Zwraca Cast No dopasowane do numeru MCT z pliku material_register.json"""
        if not mct_value or not self.material_register_data:
            return ""

        # Czyszczenie szukanego numeru MCT (usuwanie spacji, rozszerzeń, mianowników)
        clean_search = str(mct_value).strip().lower()
        if "." in clean_search:
            clean_search = clean_search.split(".")[0]
        clean_search = clean_search.replace("mct", "").strip()

        for item in self.material_register_data:
            if not isinstance(item, dict):
                continue

            # Szukamy kluczy odpowiadających MCT oraz Cast No odporne na formatowanie
            mct_match_found = False
            cast_val = ""

            for k, v in item.items():
                k_clean = k.lower().replace(" ", "").replace("_", "").replace(".", "")
                if "mct" in k_clean:
                    v_str = str(v).strip().lower()
                    if "." in v_str:
                        v_str = v_str.split(".")[0]
                    v_str = v_str.replace("mct", "").strip()
                    if clean_search and v_str == clean_search:
                        mct_match_found = True

                if "cast" in k_clean:
                    cast_val = str(v).strip()

            if mct_match_found and cast_val:
                return cast_val

        return ""

    def load_data(self):
        """Wczytuje zlecenia i konfigurację z pliku JSON bazy."""
        if os.path.exists(self.db_file):
            try:
                with open(self.db_file, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    self.clients_db = set(data.get("clients", []))
                    self.machines_db = set(data.get("machines", []))
                    self.operators_db = set(data.get("operators", []))
                    self.status_db = set(data.get("statuses", ["Planned", "In Progress", "Completed", "Delayed"]))
                    self.status_colors = data.get("status_colors", {})
                    self.mct_folder = data.get("mct_folder", "")
                    self.material_register_file = data.get("material_register_file", "")
                    self.jobs_data = data.get("jobs", [])
            except Exception as e:
                messagebox.showerror("Error", f"Failed to load database:\n{e}")
                self.jobs_data = []
        else:
            self.clients_db = {"Client A", "Client B"}
            self.machines_db = {"CNC-01", "CNC-02", "CNC-03"}
            self.operators_db = {"John Doe", "Jane Smith"}
            self.status_db = {"Planned", "In Progress", "Completed", "Delayed"}
            self.status_colors = {
                "Planned": "#e2e3e5",
                "In Progress": "#cce5ff",
                "Completed": "#d4edda",
                "Delayed": "#f8d7da"
            }
            self.mct_folder = ""
            self.material_register_file = ""
            self.jobs_data = []

        # Upewnij się, że kolory istnieją dla wszystkich statusów
        for st in self.status_db:
            if st not in self.status_colors:
                self.status_colors[st] = "#ffffff"

        self.load_material_register()

    def save_data(self):
        """Zapisuje dane do pliku bazy JSON."""
        data = {
            "clients": list(self.clients_db),
            "machines": list(self.machines_db),
            "operators": list(self.operators_db),
            "statuses": list(self.status_db),
            "status_colors": self.status_colors,
            "mct_folder": self.mct_folder,
            "material_register_file": self.material_register_file,
            "jobs": self.jobs_data
        }
        try:
            with open(self.db_file, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=4)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save database:\n{e}")

    def create_widgets(self):
        # Górny pasek narzędzi
        toolbar = tk.Frame(self.root, bg="#f0f0f0", bd=1, relief=tk.RAISED)
        toolbar.pack(side=tk.TOP, fill=tk.X)

        tk.Button(toolbar, text="➕ Add Job", bg="#e9ecef", command=self.add_main_job).pack(side=tk.LEFT, padx=5, pady=5)
        tk.Button(toolbar, text="✏️ Edit", bg="#e9ecef", command=self.edit_selected_item).pack(side=tk.LEFT, padx=5, pady=5)
        tk.Button(toolbar, text="❌ Delete", bg="#e9ecef", command=self.delete_selected_item).pack(side=tk.LEFT, padx=5, pady=5)
        
        tk.Label(toolbar, text="|", bg="#f0f0f0").pack(side=tk.LEFT, padx=5)
        tk.Button(toolbar, text="⚙️ Manage Lists", bg="#e9ecef", command=self.open_manager_dialog).pack(side=tk.LEFT, padx=5, pady=5)
        tk.Button(toolbar, text="🛠️ App Settings", bg="#e9ecef", command=self.open_settings_dialog).pack(side=tk.LEFT, padx=5, pady=5)
        tk.Button(toolbar, text="📁 MCT Folder", bg="#e9ecef", command=self.select_mct_folder).pack(side=tk.LEFT, padx=5, pady=5)
        tk.Button(toolbar, text="📜 Open MCT Cert", bg="#e9ecef", command=self.open_selected_mct).pack(side=tk.LEFT, padx=5, pady=5)

        tk.Label(toolbar, text="|", bg="#f0f0f0").pack(side=tk.LEFT, padx=5)
        tk.Button(toolbar, text="🔄 Refresh", bg="#e9ecef", command=self.refresh_tree).pack(side=tk.LEFT, padx=5, pady=5)
        tk.Button(toolbar, text="📤 Export Excel", bg="#d4edda", command=self.export_to_excel).pack(side=tk.RIGHT, padx=5, pady=5)
        tk.Button(toolbar, text="📥 Import Excel", bg="#d1ecf1", command=self.import_from_excel).pack(side=tk.RIGHT, padx=5, pady=5)

        # Panel główny z tabelą (Treeview)
        tree_frame = tk.Frame(self.root)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        columns = (
            "No", "Job / Part", "Client Name", "Due Date", "Cast No", "MCT", 
            "Drawing / File", "Sub No", "Sub Part Name", "Quantity", 
            "Machine", "Operator", "Status", "Notes"
        )
        
        self.tree = ttk.Treeview(tree_frame, columns=columns, show="headings", selectmode="extended")
        
        # Konfiguracja nagłówków i szerokości
        self.tree.heading("No", text="No")
        self.tree.heading("Job / Part", text="Job / Part Name")
        self.tree.heading("Client Name", text="Client Name 📇")
        self.tree.heading("Due Date", text="Due Date 📅")
        self.tree.heading("Cast No", text="Cast No")
        self.tree.heading("MCT", text="MCT")
        self.tree.heading("Drawing / File", text="Drawing / File")
        self.tree.heading("Sub No", text="Sub No")
        self.tree.heading("Sub Part Name", text="Sub Part Name")
        self.tree.heading("Quantity", text="Qty")
        self.tree.heading("Machine", text="Machine")
        self.tree.heading("Operator", text="Operator")
        self.tree.heading("Status", text="Status")
        self.tree.heading("Notes", text="Notes")

        col_widths = [40, 120, 110, 90, 80, 80, 110, 50, 120, 50, 80, 90, 90, 150]
        for col, width in zip(columns, col_widths):
            self.tree.column(col, width=width, anchor=tk.W)

        tree_scroll_y = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.tree.yview)
        tree_scroll_x = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscrollcommand=tree_scroll_y.set, xscrollcommand=tree_scroll_x.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        tree_scroll_y.grid(row=0, column=1, sticky="ns")
        tree_scroll_x.grid(row=1, column=0, sticky="ew")

        tree_frame.rowconfigure(0, weight=1)
        tree_frame.columnconfigure(0, weight=1)

        self.tree.bind("<Double-1>", lambda e: self.edit_selected_item())

    def open_settings_dialog(self):
        """Okno ustawień aplikacji (App Settings) – naprawiona obsługa kolorów statusów"""
        settings_win = tk.Toplevel(self.root)
        settings_win.title("App Settings")
        settings_win.geometry("480x520")
        settings_win.resizable(False, False)
        settings_win.grab_set()

        tk.Label(settings_win, text="Application Settings", font=("Arial", 12, "bold")).pack(pady=10)

        # Sekcja pliku Rejestru Materiałowego
        frame_mat = tk.LabelFrame(settings_win, text=" Material Register Settings ", font=("Arial", 10, "bold"))
        frame_mat.pack(fill=tk.X, padx=15, pady=5)

        lbl_mat_path = tk.Label(frame_mat, text=self.material_register_file if self.material_register_file else "No file selected", fg="gray", anchor="w")
        lbl_mat_path.pack(fill=tk.X, padx=10, pady=5)

        def browse_material_file():
            filename = filedialog.askopenfilename(
                title="Select Material Register JSON File",
                filetypes=[("JSON Files", "*.json"), ("All Files", "*.*")],
                parent=settings_win
            )
            if filename:
                self.material_register_file = filename
                lbl_mat_path.config(text=filename, fg="black")
                self.load_material_register()
                self.save_data()
                messagebox.showinfo("Success", "Material register file updated successfully!", parent=settings_win)

        tk.Button(frame_mat, text="📄 Select Material Register.json", command=browse_material_file).pack(anchor="w", padx=10, pady=5)

        # Sekcja Kolorów Statusów
        frame_colors = tk.LabelFrame(settings_win, text=" Status Colors Management ", font=("Arial", 10, "bold"))
        frame_colors.pack(fill=tk.BOTH, expand=True, padx=15, pady=10)

        canvas = tk.Canvas(frame_colors, borderwidth=0, highlightthickness=0)
        scroll_y = ttk.Scrollbar(frame_colors, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scroll_y.set)

        canvas.pack(side="left", fill="both", expand=True, padx=5, pady=5)
        scroll_y.pack(side="right", fill="y")

        def populate_status_color_rows():
            for widget in scrollable_frame.winfo_children():
                widget.destroy()

            # Upewnij się, że wszystkie statusy mają wpis w słowniku kolorów
            for st in sorted(self.status_db):
                if st not in self.status_colors:
                    self.status_colors[st] = "#ffffff"

            for idx, status in enumerate(sorted(self.status_db)):
                row_frame = tk.Frame(scrollable_frame)
                row_frame.pack(fill=tk.X, pady=2, padx=5)

                lbl = tk.Label(row_frame, text=status, width=18, anchor="w", font=("Arial", 9))
                lbl.pack(side=tk.LEFT, padx=5)

                current_color = self.status_colors.get(status, "#ffffff")
                
                color_box = tk.Label(row_frame, text="      ", bg=current_color, relief=tk.SOLID, bd=1)
                color_box.pack(side=tk.LEFT, padx=5)

                def pick_color(st_name, box_lbl):
                    color_code = colorchooser.askcolor(title=f"Choose color for {st_name}", parent=settings_win)
                    if color_code and color_code[1]:
                        hex_c = color_code[1]
                        self.status_colors[st_name] = hex_c
                        box_lbl.config(bg=hex_c)
                        self.save_data()
                        self.refresh_tree()

                btn_pick = tk.Button(row_frame, text="Change Color", command=lambda s=status, b=color_box: pick_color(s, b))
                btn_pick.pack(side=tk.RIGHT, padx=5)

        populate_status_color_rows()

        # Przycisk zamknięcia ustawień
        tk.Button(settings_win, text="Close", command=settings_win.destroy, width=15, bg="#e0e0e0").pack(pady=10)

    def select_mct_folder(self):
        """Wskazanie folderu z certyfikatami MCT"""
        folder = filedialog.askdirectory(title="Select MCT Certificates Folder")
        if folder:
            self.mct_folder = folder
            self.save_data()
            messagebox.showinfo("Success", f"MCT folder set to:\n{folder}")

    def open_selected_mct(self):
        """Otwieranie pliku certyfikatu MCT powiązanego z zaznaczonym wierszem"""
        selected_items = self.tree.selection()
        if not selected_items:
            messagebox.showwarning("Warning", "Please select a row with an MCT number.")
            return

        item_id = selected_items[0]
        values = self.tree.item(item_id, "values")
        # Kolumna MCT znajduje się na indeksie 5 w nowym układzie
        mct_val = values[5].strip()

        if not mct_val or mct_val == "-":
            messagebox.showwarning("Warning", "Selected item does not have an MCT number.")
            return

        if not self.mct_folder or not os.path.exists(self.mct_folder):
            messagebox.showerror("Error", "MCT certificates folder is not set or does not exist. Set it in toolbar (📁 MCT Folder).")
            return

        # Szukanie pasującego pliku w wybranym katalogu
        found_file = None
        clean_mct = mct_val.lower().replace(" ", "").replace("_", "")
        
        for root_dir, dirs, files in os.walk(self.mct_folder):
            for file in files:
                file_lower = file.lower().replace(" ", "").replace("_", "")
                if clean_mct in file_lower:
                    found_file = os.path.join(root_dir, file)
                    break
            if found_file:
                break

        if found_file:
            try:
                if os.name == 'nt':
                    os.startfile(found_file)
                elif sys.platform == 'darwin':
                    subprocess.run(['open', found_file])
                else:
                    subprocess.run(['xdg-open', found_file])
            except Exception as e:
                messagebox.showerror("Error", f"Failed to open file:\n{e}")
        else:
            messagebox.showinfo("Not Found", f"No certificate file matching MCT '{mct_val}' was found in the folder.")

    def open_manager_dialog(self):
        """Otwiera okno zarządzania listami (Klienci, Maszyny, Operatorzy, Statusy)"""
        manager_win = tk.Toplevel(self.root)
        manager_win.title("Manage Lists")
        manager_win.geometry("400x350")
        manager_win.resizable(False, False)
        manager_win.grab_set()

        tk.Label(manager_win, text="Manage Lists Database", font=("Arial", 12, "bold")).pack(pady=10)

        frame_btns = tk.Frame(manager_win)
        frame_btns.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        tk.Button(frame_btns, text="Manage Clients 📇", font=("Arial", 10), 
                  command=lambda: ListManagerDialog(manager_win, "Manage Clients", self.clients_db, self.save_data)).pack(fill=tk.X, pady=5)
        
        tk.Button(frame_btns, text="Manage Machines ⚙️", font=("Arial", 10), 
                  command=lambda: ListManagerDialog(manager_win, "Manage Machines", self.machines_db, self.save_data)).pack(fill=tk.X, pady=5)
        
        tk.Button(frame_btns, text="Manage Operators 👤", font=("Arial", 10), 
                  command=lambda: ListManagerDialog(manager_win, "Manage Operators", self.operators_db, self.save_data)).pack(fill=tk.X, pady=5)
        
        tk.Button(frame_btns, text="Manage Statuses 📌", font=("Arial", 10), 
                  command=lambda: ListManagerDialog(manager_win, "Manage Statuses", self.status_db, self.save_data)).pack(fill=tk.X, pady=5)

        tk.Button(manager_win, text="Close", width=12, command=manager_win.destroy).pack(pady=10)

    def refresh_tree(self):
        """Odświeża widok tabeli, sortując zlecenia główne rosnąco według najwcześniejszego Due Date"""
        # Zapamiętaj aktualnie zaznaczone elementy oraz otwarte gałęzie
        selected_items = self.tree.selection()
        selected_job_nos = set()
        for item_id in selected_items:
            vals = self.tree.item(item_id, "values")
            if vals:
                selected_job_nos.add(vals[0])

        open_nodes = set()
        for item_id in self.tree.get_children():
            if self.tree.item(item_id, "open"):
                vals = self.tree.item(item_id, "values")
                if vals:
                    open_nodes.add(vals[0])

        for row in self.tree.get_children():
            self.tree.delete(row)

        # Konfiguracja tagów kolorów dla statusów w Treeview
        for st, color in self.status_colors.items():
            self.tree.tag_configure(st, background=color)

        # Automatyczne sortowanie zleceń głównych według najwcześniejszego Due Date sub-elementu
        def get_earliest_due_date(job):
            sub_items = job.get("sub_items", [])
            if not sub_items:
                return "9999-12-31"
            dates = []
            for sub in sub_items:
                d_str = sub.get("due_date", "").strip()
                if d_str:
                    dates.append(d_str)
            if not dates:
                return "9999-12-31"
            return min(dates)

        sorted_jobs = sorted(self.jobs_data, key=lambda j: get_earliest_due_date(j))

        # Wstawianie danych do drzewa
        items_to_select = []
        for idx, job in enumerate(sorted_jobs, start=1):
            job_name = job.get("job_name", "")
            client_name = job.get("client_name", "")
            
            sub_items = job.get("sub_items", [])
            
            # Jeśli brak sub-elementów, wyznaczamy puste wartości dla głównego wiersza
            due_date = ""
            cast_no = ""
            mct = ""
            drawing = ""
            sub_no = ""
            sub_part_name = ""
            qty = ""
            machine = ""
            operator = ""
            status = job.get("status", "Planned")
            notes = job.get("notes", "")

            if sub_items:
                # Wiersz główny będzie rozwinięty jako parent, a sub-elementy jako dzieci
                # Wybieramy najwcześniejszą datę dla wiersza głównego
                due_date = get_earliest_due_date(job)
                if due_date == "9999-12-31":
                    due_date = ""

                parent_id = self.tree.insert(
                    "", tk.END, 
                    values=(
                        str(idx), job_name, client_name, due_date, "", "", 
                        "", "", "", "", "", "", status, notes
                    ),
                    tags=(status,)
                )

                if str(idx) in open_nodes:
                    self.tree.item(parent_id, open=True)
                if str(idx) in selected_job_nos:
                    items_to_select.append(parent_id)

                for sub in sub_items:
                    s_no = sub.get("sub_no", "")
                    s_name = sub.get("sub_part_name", "")
                    s_qty = sub.get("quantity", "")
                    s_due = sub.get("due_date", "")
                    s_mct = sub.get("mct", "")
                    s_cast = self.get_cast_no_for_mct(s_mct) # Pobieranie Cast No dla MCT
                    s_drawing = sub.get("drawing", "")
                    s_mach = sub.get("machine", "")
                    s_oper = sub.get("operator", "")
                    s_status = sub.get("status", "Planned")
                    s_notes = sub.get("notes", "")

                    child_id = self.tree.insert(
                        parent_id, tk.END,
                        values=(
                            "", "", "", s_due, s_cast, s_mct, s_drawing, 
                            s_no, s_name, s_qty, s_mach, s_oper, s_status, s_notes
                        ),
                        tags=(s_status,)
                    )
            else:
                parent_id = self.tree.insert(
                    "", tk.END,
                    values=(
                        str(idx), job_name, client_name, due_date, cast_no, mct, 
                        drawing, sub_no, sub_part_name, qty, machine, operator, status, notes
                    ),
                    tags=(status,)
                )
                if str(idx) in selected_job_nos:
                    items_to_select.append(parent_id)

        if items_to_select:
            self.tree.selection_set(items_to_select)

    def add_main_job(self):
        """Okno dialogowe dodawania nowego zlecenia głównego wraz z sub-częściami"""
        dlg = tk.Toplevel(self.root)
        dlg.title("Add New Job")
        dlg.geometry("700x600")
        dlg.grab_set()

        # Pola główne
        frame_top = tk.LabelFrame(dlg, text=" Main Job Details ", font=("Arial", 10, "bold"))
        frame_top.pack(fill=tk.X, padx=10, pady=10)

        tk.Label(frame_top, text="Job / Part Name:").grid(row=0, column=0, sticky="w", padx=5, pady=5)
        entry_job_name = tk.Entry(frame_top, width=30)
        entry_job_name.grid(row=0, column=1, sticky="w", padx=5, pady=5)

        tk.Label(frame_top, text="Client Name:").grid(row=1, column=0, sticky="w", padx=5, pady=5)
        combo_client = ttk.Combobox(frame_top, values=sorted(list(self.clients_db)), width=28)
        combo_client.grid(row=1, column=1, sticky="w", padx=5, pady=5)

        tk.Label(frame_top, text="Status:").grid(row=2, column=0, sticky="w", padx=5, pady=5)
        combo_status = ttk.Combobox(frame_top, values=sorted(list(self.status_db)), width=28, state="readonly")
        combo_status.grid(row=2, column=1, sticky="w", padx=5, pady=5)
        if self.status_db:
            combo_status.set(list(self.status_db)[0])

        tk.Label(frame_top, text="Notes:").grid(row=3, column=0, sticky="w", padx=5, pady=5)
        entry_notes = tk.Entry(frame_top, width=50)
        entry_notes.grid(row=3, column=1, sticky="w", padx=5, pady=5)

        # Sub-elementy
        frame_subs = tk.LabelFrame(dlg, text=" Sub-Parts / Operations ", font=("Arial", 10, "bold"))
        frame_subs.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        sub_columns = ("Sub No", "Sub Part Name", "Qty", "Due Date", "MCT", "Drawing", "Machine", "Operator", "Status")
        sub_tree = ttk.Treeview(frame_subs, columns=sub_columns, show="headings", height=8)
        for col in sub_columns:
            sub_tree.heading(col, text=col)
            sub_tree.column(col, width=70, anchor=tk.W)
        sub_tree.pack(side=tk.TOP, fill=tk.BOTH, expand=True, padx=5, pady=5)

        sub_items_list = []

        def add_sub_dialog():
            sub_dlg = tk.Toplevel(dlg)
            sub_dlg.title("Add Sub-Part")
            sub_dlg.geometry("350x450")
            sub_dlg.grab_set()

            tk.Label(sub_dlg, text="Sub No:").pack(anchor="w", padx=10, pady=2)
            e_sub_no = tk.Entry(sub_dlg, width=30)
            e_sub_no.pack(padx=10, pady=2)

            tk.Label(sub_dlg, text="Sub Part Name:").pack(anchor="w", padx=10, pady=2)
            e_sub_name = tk.Entry(sub_dlg, width=30)
            e_sub_name.pack(padx=10, pady=2)

            tk.Label(sub_dlg, text="Quantity:").pack(anchor="w", padx=10, pady=2)
            e_qty = tk.Entry(sub_dlg, width=30)
            e_qty.insert(0, "1")
            e_qty.pack(padx=10, pady=2)

            tk.Label(sub_dlg, text="Due Date:").pack(anchor="w", padx=10, pady=2)
            cal_due = DateEntry(sub_dlg, width=27, background='darkblue', foreground='white', date_pattern='yyyy-mm-dd')
            cal_due.pack(padx=10, pady=2)

            tk.Label(sub_dlg, text="MCT:").pack(anchor="w", padx=10, pady=2)
            e_mct = tk.Entry(sub_dlg, width=30)
            e_mct.pack(padx=10, pady=2)

            tk.Label(sub_dlg, text="Drawing / File:").pack(anchor="w", padx=10, pady=2)
            e_drawing = tk.Entry(sub_dlg, width=30)
            e_drawing.pack(padx=10, pady=2)

            tk.Label(sub_dlg, text="Machine:").pack(anchor="w", padx=10, pady=2)
            c_mach = ttk.Combobox(sub_dlg, values=sorted(list(self.machines_db)), width=28)
            c_mach.pack(padx=10, pady=2)

            tk.Label(sub_dlg, text="Operator:").pack(anchor="w", padx=10, pady=2)
            c_oper = ttk.Combobox(sub_dlg, values=sorted(list(self.operators_db)), width=28)
            c_oper.pack(padx=10, pady=2)

            def save_sub():
                s_data = {
                    "sub_no": e_sub_no.get().strip(),
                    "sub_part_name": e_sub_name.get().strip(),
                    "quantity": e_qty.get().strip(),
                    "due_date": cal_due.get(),
                    "mct": e_mct.get().strip(),
                    "drawing": e_drawing.get().strip(),
                    "machine": c_mach.get().strip(),
                    "operator": c_oper.get().strip(),
                    "status": combo_status.get()
                }
                sub_items_list.append(s_data)
                sub_tree.insert("", tk.END, values=(
                    s_data["sub_no"], s_data["sub_part_name"], s_data["quantity"],
                    s_data["due_date"], s_data["mct"], s_data["drawing"],
                    s_data["machine"], s_data["operator"], s_data["status"]
                ))
                sub_dlg.destroy()

            tk.Button(sub_dlg, text="Add Sub-Part", bg="#d4edda", command=save_sub, width=15).pack(pady=15)

        def remove_sub():
            selected = sub_tree.selection()
            if selected:
                idx = sub_tree.index(selected[0])
                sub_tree.delete(selected[0])
                del sub_items_list[idx]

        btn_f = tk.Frame(frame_subs)
        btn_f.pack(fill=tk.X, padx=5, pady=5)
        tk.Button(btn_f, text="➕ Add Sub-Part", command=add_sub_dialog).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_f, text="❌ Remove Sub-Part", bg="#f8d7da", command=remove_sub).pack(side=tk.LEFT, padx=5)

        def save_job():
            j_name = entry_job_name.get().strip()
            c_name = combo_client.get().strip()
            if not j_name:
                messagebox.showwarning("Warning", "Job / Part Name is required!", parent=dlg)
                return

            if c_name:
                self.clients_db.add(c_name)

            new_job = {
                "job_name": j_name,
                "client_name": c_name,
                "status": combo_status.get(),
                "notes": entry_notes.get().strip(),
                "sub_items": sub_items_list
            }

            self.jobs_data.append(new_job)
            self.save_data()
            self.refresh_tree()
            dlg.destroy()

        tk.Button(dlg, text="Save Job", bg="#d4edda", font=("Arial", 10, "bold"), width=15, command=save_job).pack(pady=15)

    def edit_selected_item(self):
        """Edycja wybranego zlecenia lub sub-części"""
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select a row to edit.")
            return

        item_id = selected[0]
        values = self.tree.item(item_id, "values")
        parent_id = self.tree.parent(item_id)

        if not parent_id:
            # Edycja głównego zlecenia
            job_no_str = values[0]
            try:
                job_idx = int(job_no_str) - 1
            except ValueError:
                return

            job = self.jobs_data[job_idx]

            dlg = tk.Toplevel(self.root)
            dlg.title("Edit Main Job")
            dlg.geometry("500x300")
            dlg.grab_set()

            tk.Label(dlg, text="Job / Part Name:").pack(anchor="w", padx=20, pady=5)
            e_job = tk.Entry(dlg, width=40)
            e_job.insert(0, job.get("job_name", ""))
            e_job.pack(anchor="w", padx=20, pady=2)

            tk.Label(dlg, text="Client Name:").pack(anchor="w", padx=20, pady=5)
            c_client = ttk.Combobox(dlg, values=sorted(list(self.clients_db)), width=38)
            c_client.set(job.get("client_name", ""))
            c_client.pack(anchor="w", padx=20, pady=2)

            tk.Label(dlg, text="Status:").pack(anchor="w", padx=20, pady=5)
            c_status = ttk.Combobox(dlg, values=sorted(list(self.status_db)), width=38, state="readonly")
            c_status.set(job.get("status", "Planned"))
            c_status.pack(anchor="w", padx=20, pady=2)

            tk.Label(dlg, text="Notes:").pack(anchor="w", padx=20, pady=5)
            e_notes = tk.Entry(dlg, width=40)
            e_notes.insert(0, job.get("notes", ""))
            e_notes.pack(anchor="w", padx=20, pady=2)

            def save_edit():
                job["job_name"] = e_job.get().strip()
                cl_val = c_client.get().strip()
                job["client_name"] = cl_val
                if cl_val:
                    self.clients_db.add(cl_val)
                job["status"] = c_status.get()
                job["notes"] = e_notes.get().strip()

                self.save_data()
                self.refresh_tree()
                dlg.destroy()

            tk.Button(dlg, text="Save Changes", bg="#d4edda", command=save_edit, width=15).pack(pady=20)

        else:
            # Edycja sub-części
            parent_values = self.tree.item(parent_id, "values")
            parent_job_idx = int(parent_values[0]) - 1
            job = self.jobs_data[parent_job_idx]

            # Znajdź sub-element na podstawie numeru sub oraz danych
            sub_no_val = values[7]
            sub_idx = -1
            for idx, s in enumerate(job.get("sub_items", [])):
                if str(s.get("sub_no")) == str(sub_no_val):
                    sub_idx = idx
                    break

            if sub_idx == -1:
                return

            sub = job["sub_items"][sub_idx]

            dlg = tk.Toplevel(self.root)
            dlg.title("Edit Sub-Part")
            dlg.geometry("350x450")
            dlg.grab_set()

            tk.Label(dlg, text="Sub No:").pack(anchor="w", padx=10, pady=2)
            e_sub_no = tk.Entry(dlg, width=30)
            e_sub_no.insert(0, sub.get("sub_no", ""))
            e_sub_no.pack(padx=10, pady=2)

            tk.Label(dlg, text="Sub Part Name:").pack(anchor="w", padx=10, pady=2)
            e_sub_name = tk.Entry(dlg, width=30)
            e_sub_name.insert(0, sub.get("sub_part_name", ""))
            e_sub_name.pack(padx=10, pady=2)

            tk.Label(dlg, text="Quantity:").pack(anchor="w", padx=10, pady=2)
            e_qty = tk.Entry(dlg, width=30)
            e_qty.insert(0, sub.get("quantity", ""))
            e_qty.pack(padx=10, pady=2)

            tk.Label(dlg, text="Due Date:").pack(anchor="w", padx=10, pady=2)
            cal_due = DateEntry(dlg, width=27, background='darkblue', foreground='white', date_pattern='yyyy-mm-dd')
            try:
                cal_due.set_date(sub.get("due_date", datetime.now().strftime('%Y-%m-%d')))
            except Exception:
                pass
            cal_due.pack(padx=10, pady=2)

            tk.Label(dlg, text="MCT:").pack(anchor="w", padx=10, pady=2)
            e_mct = tk.Entry(dlg, width=30)
            e_mct.insert(0, sub.get("mct", ""))
            e_mct.pack(padx=10, pady=2)

            tk.Label(dlg, text="Drawing / File:").pack(anchor="w", padx=10, pady=2)
            e_drawing = tk.Entry(dlg, width=30)
            e_drawing.insert(0, sub.get("drawing", ""))
            e_drawing.pack(padx=10, pady=2)

            tk.Label(dlg, text="Machine:").pack(anchor="w", padx=10, pady=2)
            c_mach = ttk.Combobox(dlg, values=sorted(list(self.machines_db)), width=28)
            c_mach.set(sub.get("machine", ""))
            c_mach.pack(padx=10, pady=2)

            tk.Label(dlg, text="Operator:").pack(anchor="w", padx=10, pady=2)
            c_oper = ttk.Combobox(dlg, values=sorted(list(self.operators_db)), width=28)
            c_oper.set(sub.get("operator", ""))
            c_oper.pack(padx=10, pady=2)

            tk.Label(dlg, text="Status:").pack(anchor="w", padx=10, pady=2)
            c_status = ttk.Combobox(dlg, values=sorted(list(self.status_db)), width=28, state="readonly")
            c_status.set(sub.get("status", "Planned"))
            c_status.pack(padx=10, pady=2)

            def save_sub_edit():
                sub["sub_no"] = e_sub_no.get().strip()
                sub["sub_part_name"] = e_sub_name.get().strip()
                sub["quantity"] = e_qty.get().strip()
                sub["due_date"] = cal_due.get()
                sub["mct"] = e_mct.get().strip()
                sub["drawing"] = e_drawing.get().strip()
                mach_val = c_mach.get().strip()
                oper_val = c_oper.get().strip()
                sub["machine"] = mach_val
                sub["operator"] = oper_val
                sub["status"] = c_status.get()

                if mach_val:
                    self.machines_db.add(mach_val)
                if oper_val:
                    self.operators_db.add(oper_val)

                self.save_data()
                self.refresh_tree()
                dlg.destroy()

            tk.Button(dlg, text="Save Sub-Part", bg="#d4edda", command=save_sub_edit, width=15).pack(pady=15)

    def delete_selected_item(self):
        """Usuwanie zaznaczonego zlecenia głównego lub sub-części"""
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select a row to delete.")
            return

        if not messagebox.askyesno("Confirmation", "Are you sure you want to delete selected item(s)?"):
            return

        for item_id in selected:
            values = self.tree.item(item_id, "values")
            parent_id = self.tree.parent(item_id)

            if not parent_id:
                # Usunięcie głównego zlecenia
                job_no_str = values[0]
                try:
                    job_idx = int(job_no_str) - 1
                    del self.jobs_data[job_idx]
                except Exception:
                    pass
            else:
                # Usunięcie sub-części
                parent_values = self.tree.item(parent_id, "values")
                parent_job_idx = int(parent_values[0]) - 1
                job = self.jobs_data[parent_job_idx]
                sub_no_val = values[7]
                
                job["sub_items"] = [s for s in job.get("sub_items", []) if str(s.get("sub_no")) != str(sub_no_val)]

        self.save_data()
        self.refresh_tree()

    def export_to_excel(self):
        """Eksport danych do pliku Excel (.xlsx)"""
        if Workbook is None:
            messagebox.showerror("Error", "openpyxl library is not installed. Cannot export to Excel.")
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Export to Excel"
        )
        if not file_path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "CNC Schedule"

            headers = [
                "No", "Job / Part Name", "Client Name", "Due Date", "Cast No", "MCT", 
                "Drawing / File", "Sub No", "Sub Part Name", "Quantity", 
                "Machine", "Operator", "Status", "Notes"
            ]
            ws.append(headers)

            # Stylowanie nagłówka
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            row_idx = 2
            for idx, job in enumerate(self.jobs_data, start=1):
                j_name = job.get("job_name", "")
                c_name = job.get("client_name", "")
                notes = job.get("notes", "")
                status = job.get("status", "Planned")

                sub_items = job.get("sub_items", [])
                if sub_items:
                    for sub in sub_items:
                        s_no = sub.get("sub_no", "")
                        s_name = sub.get("sub_part_name", "")
                        s_qty = sub.get("quantity", "")
                        s_due = sub.get("due_date", "")
                        s_mct = sub.get("mct", "")
                        s_cast = self.get_cast_no_for_mct(s_mct)
                        s_drawing = sub.get("drawing", "")
                        s_mach = sub.get("machine", "")
                        s_oper = sub.get("operator", "")
                        s_status = sub.get("status", status)

                        ws.append([str(idx), j_name, c_name, s_due, s_cast, s_mct, s_drawing, s_no, s_name, s_qty, s_mach, s_oper, s_status, notes])
                        row_idx += 1
                else:
                    ws.append([str(idx), j_name, c_name, "", "", "", "", "", "", "", "", "", status, notes])
                    row_idx += 1

            wb.save(file_path)
            messagebox.showinfo("Success", f"Data successfully exported to:\n{file_path}")
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export Excel file:\n{str(e)}")

    def import_from_excel(self):
        """Import danych z pliku Excel (.xlsx)"""
        if load_workbook is None:
            messagebox.showerror("Error", "openpyxl library is not installed. Cannot import from Excel.")
            return

        file_path = filedialog.askopenfilename(
            title="Import from Excel",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if not file_path:
            return

        try:
            wb = load_workbook(file_path)
            ws = wb.active
            imported_rows = 0

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue

                # Oczekiwana kolejność kolumn w pliku Excel
                # 0: No, 1: Job Name, 2: Client, 3: Due Date, 4: Cast No, 5: MCT, 6: Drawing, 7: Sub No, 8: Sub Name, 9: Qty, 10: Machine, 11: Operator, 12: Status, 13: Notes
                job_name = str(row[1]) if row[1] is not None else "Unknown Job"
                client = str(row[2]) if row[2] is not None else ""
                due_date = str(row[3]) if row[3] is not None else ""
                mct = str(row[5]) if row[5] is not None else ""
                drawing = str(row[6]) if row[6] is not None else ""
                sub_no = str(row[7]) if row[7] is not None else ""
                sub_name = str(row[8]) if row[8] is not None else ""
                qty = str(row[9]) if row[9] is not None else "1"
                machine = str(row[10]) if row[10] is not None else ""
                operator = str(row[11]) if row[11] is not None else ""
                status = str(row[12]) if row[12] is not None else "Planned"
                notes = str(row[13]) if row[13] is not None else ""

                # Szukaj czy zlecenie główne już istnieje
                target_job = None
                for j in self.jobs_data:
                    if j.get("job_name") == job_name:
                        target_job = j
                        break

                if not target_job:
                    target_job = {
                        "job_name": job_name,
                        "client_name": client,
                        "status": status,
                        "notes": notes,
                        "sub_items": []
                    }
                    self.jobs_data.append(target_job)

                sub_item = {
                    "sub_no": sub_no,
                    "sub_part_name": sub_name,
                    "quantity": qty,
                    "due_date": due_date,
                    "mct": mct,
                    "drawing": drawing,
                    "machine": machine,
                    "operator": operator,
                    "status": status,
                    "notes": notes
                }

                if sub_no:
                    existing_subs = {s["sub_no"]: s for s in target_job.get("sub_items", [])}
                    if sub_item["sub_no"] in existing_subs:
                        existing_subs[sub_item["sub_no"]].update(sub_item)
                    else:
                        target_job.setdefault("sub_items", []).append(sub_item)

                if client:
                    self.clients_db.add(client)
                if machine:
                    self.machines_db.add(machine)
                if operator:
                    self.operators_db.add(operator)
                if status:
                    self.status_db.add(status)
                    if status not in self.status_colors:
                        self.status_colors[status] = "#ffffff"

                imported_rows += 1

            self.save_data()
            self.refresh_tree()
            messagebox.showinfo("Import Successful", f"Imported {imported_rows} rows from Excel file.")
        except Exception as e:
            messagebox.showerror("Import Error", f"Failed to import Excel data:\n{str(e)}")


if __name__ == "__main__":
    root = tk.Tk()
    root.state("zoomed")

    # Plansza informacyjna (Splash Screen) na start
    splash = tk.Toplevel(root)
    splash.overrideredirect(True)
    splash.geometry("400x200")
    splash.config(bg="#1f4e78")
    
    # Wyśrodkowanie okna splash screen
    splash.update_idletasks()
    sw = splash.winfo_screenwidth()
    sh = splash.winfo_screenheight()
    x = (sw - 400) // 2
    y = (sh - 200) // 2
    splash.geometry(f"400x200+{x}+{y}")

    tk.Label(splash, text="CNC Production Schedule", font=("Arial", 16, "bold"), fg="white", bg="#1f4e78").pack(pady=(40, 10))
    tk.Label(splash, text="Loading application database and settings...", font=("Arial", 10), fg="#d0d0d0", bg="#1f4e78").pack(pady=10)
    
    progress = ttk.Progressbar(splash, orient="horizontal", length=300, mode="indeterminate")
    progress.pack(pady=10)
    progress.start(10)

    root.withdraw()

    def launch_main():
        splash.destroy()
        root.deiconify()
        ScheduleApp(root)

    root.after(3000, launch_main)
    root.mainloop()