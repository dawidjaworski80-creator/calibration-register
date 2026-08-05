import csv
import json
import os
import subprocess
import sys
import tempfile
import tkinter as tk
from datetime import datetime, timedelta
from tkinter import filedialog, messagebox, ttk, colorchooser

try:
    from tkcalendar import DateEntry
except ImportError:
    DateEntry = None

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    Workbook = None
    load_workbook = None
    Alignment = None
    Font = None
    PatternFill = None

if DateEntry is None:
    class DateEntry(tk.Entry):
        def __init__(self, parent, *args, **kwargs):
            self.date_pattern = kwargs.pop("date_pattern", "dd-mm-yyyy")
            self.locale = kwargs.pop("locale", None)
            super().__init__(parent, *args, **kwargs)
            self.insert(0, datetime.now().strftime("%d-%m-%Y"))

        def get_date(self):
            value = self.get().strip()
            for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y"):
                try:
                    return datetime.strptime(value, fmt).date()
                except Exception:
                    continue
            return datetime.now().date()

        def set_date(self, date_value):
            if date_value is None:
                return
            try:
                if isinstance(date_value, datetime):
                    date_value = date_value.date()
                self.delete(0, tk.END)
                self.insert(0, date_value.strftime("%d-%m-%Y"))
            except Exception:
                pass

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    Workbook = None
    Alignment = None
    Font = None
    PatternFill = None

# Sprawdzenie ścieżki uruchomienia aplikacji
if getattr(sys, "frozen", False):
    APP_DIR = os.path.dirname(sys.executable)
else:
    APP_DIR = os.path.dirname(os.path.abspath(__file__))

CONFIG_FILE = os.path.join(APP_DIR, "config.json")
DEFAULT_DB_FILE = os.path.join(APP_DIR, "cnc_schedule_db.json")


def get_saved_db_path():
    """Pobiera ścieżkę do bazy z config.json lub zwraca domyślną."""
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                cfg = json.load(f)
                path = cfg.get("db_file_path", DEFAULT_DB_FILE)
                if path:
                    return path
        except Exception:
            pass
    return DEFAULT_DB_FILE


def save_db_path(path):
    """Zapisuje ścieżkę do bazy w pliku config.json."""
    try:
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump({"db_file_path": path}, f, indent=4)
    except Exception:
        pass


class ListManagerDialog(tk.Toplevel):
    """Uniwersalne okno dialogowe do zarządzania listami (Klienci, Maszyny, Operatorzy, Statusy)"""

    def __init__(self, parent, title_name, item_set, callback_on_update, color_map=None):
        super().__init__(parent)
        self.title(f"Manage {title_name}")
        self.geometry("360x320")
        self.resizable(False, False)
        self.grab_set()

        self.title_name = title_name
        self.item_set = item_set
        self.callback_on_update = callback_on_update
        self.color_map = color_map or {}

        tk.Label(
            self, text=f"Manage {title_name} List", font=("Arial", 11, "bold")
        ).pack(pady=8)

        input_frame = tk.Frame(self)
        input_frame.pack(fill="x", padx=10, pady=5)

        self.entry_item = tk.Entry(input_frame, width=20)
        self.entry_item.pack(side="left", padx=(0, 5), expand=True, fill="x")

        # Color selector for Statuses
        if self.title_name.lower() == "statuses":
            self.color_entry = tk.Entry(input_frame, width=10)
            self.color_entry.pack(side="left", padx=(4, 0))
            tk.Button(
                input_frame,
                text="Pick",
                font=("Arial", 7),
                command=self.pick_color,
            ).pack(side="left", padx=(2, 0))
            self.color_preview = tk.Label(
                input_frame,
                width=2,
                bg="#ffffff",
                relief="solid",
                borderwidth=1,
            )
            self.color_preview.pack(side="left", padx=(4, 0))

        btn_add = tk.Button(
            input_frame,
            text="Add",
            bg="#27ae60",
            fg="white",
            font=("Arial", 8, "bold"),
            command=self.add_item,
        )
        btn_add.pack(side="right")

        # Update Color button for statuses
        if self.title_name.lower() == "statuses":
            btn_update = tk.Button(
                input_frame,
                text="Update Color",
                bg="#f39c12",
                fg="white",
                font=("Arial", 8, "bold"),
                command=self.update_color,
            )
            btn_update.pack(side="right", padx=(4, 2))

        self.listbox = tk.Listbox(self, font=("Arial", 9))
        self.listbox.pack(fill="both", expand=True, padx=10, pady=5)

        btn_del = tk.Button(
            self,
            text=f"Delete Selected {title_name}",
            bg="#c0392b",
            fg="white",
            font=("Arial", 9, "bold"),
            command=self.delete_item,
        )
        btn_del.pack(fill="x", padx=10, pady=(0, 10))

        self.listbox.bind("<<ListboxSelect>>", self.on_select)
        self.refresh_list()

    def refresh_list(self):
        self.listbox.delete(0, tk.END)
        for item in sorted(self.item_set):
            self.listbox.insert(tk.END, item)

    def pick_color(self):
        initial_color = self.color_entry.get().strip() or "#ffffff"
        color_code = colorchooser.askcolor(initialcolor=initial_color, title="Pick Status Color")
        if color_code and color_code[1]:
            self.color_entry.delete(0, tk.END)
            self.color_entry.insert(0, color_code[1])
            if hasattr(self, "color_preview"):
                self.color_preview.config(bg=color_code[1])

    def add_item(self):
        name = self.entry_item.get().strip()
        if not name:
            messagebox.showwarning("Warning", "Please enter a value to add.")
            return
        if name in self.item_set:
            messagebox.showwarning("Warning", f"'{name}' already exists in {self.title_name}.")
            return
        self.item_set.add(name)
        if self.title_name.lower() == "statuses":
            self.color_map[name] = self.color_entry.get().strip() or "#ffffff"
        self.refresh_list()
        self.callback_on_update()

    def delete_item(self):
        sel = self.listbox.curselection()
        if not sel:
            return
        name = self.listbox.get(sel[0])
        if messagebox.askyesno(
            "Confirm Delete",
            f"Delete '{name}' from {self.title_name}?",
        ):
            self.item_set.discard(name)
            if self.title_name.lower() == "statuses":
                self.color_map.pop(name, None)
            self.refresh_list()
            self.callback_on_update()

    def update_color(self):
        if self.title_name.lower() != "statuses":
            return
        sel = self.listbox.curselection()
        if not sel:
            messagebox.showwarning("Warning", "Please select a status to update the color.")
            return
        name = self.listbox.get(sel[0])
        color = self.color_entry.get().strip() or "#ffffff"
        self.color_map[name] = color
        self.refresh_list()
        self.callback_on_update()

    def on_select(self, event):
        sel = self.listbox.curselection()
        if not sel:
            return
        name = self.listbox.get(sel[0])
        self.entry_item.delete(0, tk.END)
        self.entry_item.insert(0, name)
        if self.title_name.lower() == "statuses":
            color = self.color_map.get(name, "")
            self.color_entry.delete(0, tk.END)
            self.color_entry.insert(0, color)
            if hasattr(self, "color_preview"):
                self.color_preview.config(bg=color if color else "#ffffff")


class DataManager:
    """Zarządza danymi aplikacji, ładowniem/zapisem JSON oraz filtracją danych."""

    def __init__(self, db_file_path):
        self.db_file_path = db_file_path
        self.clients_db = set()
        self.machines_db = set()
        self.operators_db = set()
        self.status_db = {
            "In Progress",
            "Sub Con",
            "Finished",
            "Inspection",
            "On Hold",
            "Ready For Milling",
        }
        # Mapping status -> color hex (strings). Keys are case-sensitive as displayed.
        self.status_colors = {
            "In Progress": "#e7f3ff",
            "Sub Con": "#fff3cd",
            "Finished": "#d4edda",
            "Inspection": "#f0e6ff",
            "On Hold": "#f5c16c",
            "Ready For Milling": "#e6fffa",
        }
        self.all_jobs_data = {}
        self.mct_dir = ""
        self.last_db_mtime = 0

    def load_data(self):
        if os.path.exists(self.db_file_path):
            try:
                with open(self.db_file_path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    self.clients_db.update(data.get("clients", []))
                    self.machines_db.update(data.get("machines", []))
                    self.operators_db.update(data.get("operators", []))
                    # support both old list-of-statuses and new dict-of-color mappings
                    if "statuses" in data:
                        s = data["statuses"]
                        if isinstance(s, dict):
                            # statuses stored as {name: color}
                            self.status_db.update(s.keys())
                            # merge colors
                            for k, v in s.items():
                                self.status_colors[k] = v
                        else:
                            self.status_db.update(s)
                    # legacy separate status_colors key
                    if "status_colors" in data and isinstance(data["status_colors"], dict):
                        for k, v in data["status_colors"].items():
                            self.status_colors[k] = v
                    self.all_jobs_data = data.get("jobs", {})
                    self._normalize_dates_in_data()
                    self.mct_dir = data.get("mct_dir", "")
                    self.last_db_mtime = os.path.getmtime(self.db_file_path)
            except Exception as e:
                raise RuntimeError(f"Failed to load database file:\n{str(e)}")

    def _normalize_dates_in_data(self):
        def parse_date(value):
            if not value:
                return None
            for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y"):
                try:
                    return datetime.strptime(value, fmt).date()
                except Exception:
                    continue
            return None

        for job_data in self.all_jobs_data.values():
            for item in job_data.get("sub_items", []):
                raw_date = item.get("date", "")
                parsed = parse_date(raw_date)
                if parsed:
                    item["date"] = parsed.strftime("%d-%m-%Y")

    def save_data(self):
        """Trwały i bezpieczny (atomiczny) zapis do pliku bazy JSON."""
        data = {
            "clients": sorted(list(self.clients_db)),
            "machines": sorted(list(self.machines_db)),
            "operators": sorted(list(self.operators_db)),
            # store statuses as dict mapping -> color so colors persist
            "statuses": {s: self.status_colors.get(s, "") for s in sorted(list(self.status_db))},
            "mct_dir": self.mct_dir,
            "jobs": self.all_jobs_data,
        }
        try:
            dir_name = os.path.dirname(self.db_file_path) or "."
            os.makedirs(dir_name, exist_ok=True)

            with tempfile.NamedTemporaryFile("w", dir=dir_name, delete=False, encoding="utf-8") as tf:
                json.dump(data, tf, indent=4, ensure_ascii=False)
                temp_name = tf.name

            os.replace(temp_name, self.db_file_path)
            self.last_db_mtime = os.path.getmtime(self.db_file_path)
        except Exception as e:
            raise RuntimeError(f"Failed to save database file:\n{str(e)}")


class SplashScreen(tk.Toplevel):
    """Plansza informacyjna (Splash Screen) wyświetlana przed główną aplikacją."""

    def __init__(self, parent, delay_ms=3000):
        super().__init__(parent)
        self.parent = parent

        # Window title
        self.title("Machine Shop Schedule Pro - Start")

        # Rozmiar okna
        width = 500
        height = 280

        # Wyśrodkowanie okna na ekranie
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        x = (screen_width // 2) - (width // 2)
        y = (screen_height // 2) - (height // 2)
        self.geometry(f"{width}x{height}+{x}+{y}")

        # Usunięcie obramowania okna
        self.overrideredirect(True)

        # Stylizacja i tło okna
        self.configure(
            bg="#2c3e50", highlightbackground="#34495e", highlightthickness=2
        )

        # Zawartość planszy informacyjnej
        title_label = tk.Label(
            self,
            text="Machine Shop Schedule Pro",
            font=("Arial", 18, "bold"),
            fg="#ffffff",
            bg="#2c3e50",
        )
        title_label.pack(pady=(40, 10))

        subtitle_label = tk.Label(
            self,
            text="Production Scheduling Management System",
            font=("Arial", 11, "italic"),
            fg="#bdc3c7",
            bg="#2c3e50",
        )
        subtitle_label.pack(pady=(0, 20))

        info_label = tk.Label(
            self,
            text="Version 4.2 | Loading resources...",
            font=("Arial", 9),
            fg="#ecf0f1",
            bg="#2c3e50",
        )
        info_label.pack(pady=(10, 0))

        self.update_idletasks()
        self.lift()
        self.attributes("-topmost", True)
        self.after(50, lambda: self.attributes("-topmost", False))

        # Zaplanowanie zamknięcia planszy i wywołania głównego okna
        self.after(delay_ms, self.close_splash)

    def close_splash(self):
        self.destroy()
        self.parent.deiconify()  # Pokazuje główne okno aplikacji


class ScheduleApp:

    def __init__(self, root):
        self.root = root
        self.root.title("Machine Shop Schedule Pro - Job Sub-parts /1 /2 /3")
        self.root.state("zoomed")
        self.root.minsize(1200, 700)
        self.root.configure(bg="#f4f4f9")

        # Inicjalizacja menedżera danych
        db_path = get_saved_db_path()
        self.data_mgr = DataManager(db_path)

        self.priorities_db = ["Low", "Normal", "High", "URGENT"]

        self.editing_job_id = None
        self.editing_sub_idx = None
        self.temp_sub_items = []
        self._reminded_dates = {}
        
        self.sort_column = "date"
        self.sort_reverse = False

        # --- HEADER ---
        header_frame = tk.Frame(self.root, bg="#2c3e50", pady=12)
        header_frame.pack(fill="x")

        title_label = tk.Label(
            header_frame,
            text="Machine Shop Schedule Pro",
            font=("Arial", 16, "bold"),
            fg="white",
            bg="#2c3e50",
        )
        title_label.pack(side="left", padx=20)

        self.clock_label = tk.Label(
            header_frame,
            text="",
            font=("Arial", 11, "bold"),
            fg="#ecf0f1",
            bg="#2c3e50",
        )
        self.clock_label.pack(side="right", padx=20)

        # --- TOP TOOLBAR & SEARCH / FILTER BAR ---
        top_bar = tk.Frame(self.root, bg="#e2e8f0", pady=8, padx=15)
        top_bar.pack(fill="x")

        self.btn_toggle_form = tk.Button(
            top_bar,
            text="+ Add New Job",
            bg="#27ae60",
            fg="white",
            font=("Arial", 9, "bold"),
            padx=10,
            pady=4,
            command=self.open_add_form,
        )
        self.btn_toggle_form.pack(side="left", padx=(0, 10))

        btn_refresh = tk.Button(
            top_bar,
            text="🔄 Refresh",
            bg="#2980b9",
            fg="white",
            font=("Arial", 9, "bold"),
            padx=8,
            pady=4,
            command=self.load_data,
        )
        btn_refresh.pack(side="left", padx=(0, 5))

        btn_reminders = tk.Button(
            top_bar,
            text="🔔 Reminders",
            bg="#d35400",
            fg="white",
            font=("Arial", 9, "bold"),
            padx=8,
            pady=4,
            command=self.open_reminder_window,
        )
        btn_reminders.pack(side="left", padx=(0, 5))

        btn_app_settings = tk.Button(
            top_bar,
            text="⚙️ App Settings",
            bg="#34495e",
            fg="white",
            font=("Arial", 9, "bold"),
            padx=8,
            pady=4,
            command=self.open_app_settings_menu,
        )
        btn_app_settings.pack(side="left", padx=(0, 15))

        tk.Label(
            top_bar, text="Search:", font=("Arial", 9, "bold"), bg="#e2e8f0"
        ).pack(side="left", padx=(5, 2))
        self.search_var = tk.StringVar()
        self.search_var.trace_add("write", lambda *args: self.apply_filters())
        self.entry_search = tk.Entry(
            top_bar, textvariable=self.search_var, width=18, font=("Arial", 9)
        )
        self.entry_search.pack(side="left", padx=(0, 15))

        tk.Label(
            top_bar, text="Filter Machine:", font=("Arial", 9, "bold"), bg="#e2e8f0"
        ).pack(side="left", padx=(5, 2))
        self.filter_machine_var = tk.StringVar(value="All")
        self.combo_filter_machine = ttk.Combobox(
            top_bar,
            textvariable=self.filter_machine_var,
            state="readonly",
            width=15,
        )
        self.combo_filter_machine.pack(side="left", padx=(0, 15))
        self.combo_filter_machine.bind(
            "<<ComboboxSelected>>", lambda e: self.apply_filters()
        )

        tk.Label(
            top_bar, text="Filter Status:", font=("Arial", 9, "bold"), bg="#e2e8f0"
        ).pack(side="left", padx=(5, 2))
        self.filter_status_var = tk.StringVar(value="All")
        self.combo_filter_status = ttk.Combobox(
            top_bar, textvariable=self.filter_status_var, state="readonly", width=12
        )
        self.combo_filter_status.pack(side="left", padx=(0, 10))
        self.combo_filter_status.bind(
            "<<ComboboxSelected>>", lambda e: self.apply_filters()
        )

        tk.Label(
            top_bar, text="Filter Client:", font=("Arial", 9, "bold"), bg="#e2e8f0"
        ).pack(side="left", padx=(5, 2))
        self.filter_client_var = tk.StringVar(value="All")
        self.combo_filter_client = ttk.Combobox(
            top_bar, textvariable=self.filter_client_var, state="readonly", width=16
        )
        self.combo_filter_client.pack(side="left", padx=(0, 10))
        self.combo_filter_client.bind(
            "<<ComboboxSelected>>", lambda e: self.apply_filters()
        )

        tk.Label(
            top_bar, text="Due Date:", font=("Arial", 9, "bold"), bg="#e2e8f0"
        ).pack(side="left", padx=(5, 2))
        self.filter_date_var = tk.StringVar()
        self.entry_filter_date = tk.Entry(
            top_bar, textvariable=self.filter_date_var, width=10, font=("Arial", 9)
        )
        self.entry_filter_date.pack(side="left", padx=(0, 10))
        self.entry_filter_date.bind("<Return>", lambda e: self.apply_filters())

        btn_reset_filters = tk.Button(
            top_bar,
            text="Reset Filters",
            bg="#7f8c8d",
            fg="white",
            font=("Arial", 8, "bold"),
            command=self.reset_filters,
        )
        btn_reset_filters.pack(side="left")

        btn_import = tk.Button(
            top_bar,
            text="📥 Import from Excel",
            bg="#16a085",
            fg="white",
            font=("Arial", 9, "bold"),
            command=self.import_from_excel,
        )
        btn_import.pack(side="right", padx=(0, 5))

        btn_export = tk.Button(
            top_bar,
            text="📊 Export to Excel",
            bg="#8e44ad",
            fg="white",
            font=("Arial", 9, "bold"),
            command=self.export_to_excel,
        )
        btn_export.pack(side="right")

        # --- FORMULARZ DODAWANIA/EDYCJI ---
        self.form_frame = tk.LabelFrame(
            self.root,
            text=" Add New Job & Sub-parts ",
            font=("Arial", 10, "bold"),
            padx=10,
            pady=8,
            bg="#f4f4f9",
        )

        header_inputs = tk.Frame(self.form_frame, bg="#f4f4f9")
        header_inputs.pack(fill="x", pady=(0, 5))

        tk.Label(
            header_inputs,
            text="Client Name:",
            font=("Arial", 9, "bold"),
            bg="#f4f4f9",
        ).grid(row=0, column=0, sticky="w", padx=4)
        client_container = tk.Frame(header_inputs, bg="#f4f4f9")
        client_container.grid(row=0, column=1, padx=4, sticky="w")
        self.combo_client = ttk.Combobox(client_container, width=16)
        self.combo_client.pack(side="left")
        tk.Button(
            client_container,
            text="📇",
            font=("Arial", 7),
            command=lambda: self.open_list_manager("Clients", self.data_mgr.clients_db),
        ).pack(side="left", padx=(2, 0))

        tk.Label(
            header_inputs,
            text="Job No. (e.g. 1234):",
            font=("Arial", 9, "bold"),
            bg="#f4f4f9",
        ).grid(row=0, column=2, sticky="w", padx=4)
        self.entry_job = tk.Entry(header_inputs, width=14)
        self.entry_job.grid(row=0, column=3, padx=4, sticky="w")
        self.entry_job.bind(
            "<KeyRelease>", lambda e: self.update_sub_item_numbers()
        )

        tk.Label(
            header_inputs, text="PO Number:", font=("Arial", 9, "bold"), bg="#f4f4f9"
        ).grid(row=0, column=4, sticky="w", padx=4)
        self.entry_po = tk.Entry(header_inputs, width=14)
        self.entry_po.grid(row=0, column=5, padx=4, sticky="w")

        sub_frame = tk.LabelFrame(
            self.form_frame,
            text=" Job Sub-parts (Numbered automatically as JobNo/1, JobNo/2, ...) ",
            font=("Arial", 9, "bold"),
            bg="#eef2f7",
            padx=8,
            pady=5,
        )
        sub_frame.pack(fill="x", pady=5)

        sub_inputs = tk.Frame(sub_frame, bg="#eef2f7")
        sub_inputs.pack(fill="x", pady=2)

        tk.Label(sub_inputs, text="Part Name:", bg="#eef2f7").grid(
            row=0, column=0, sticky="w"
        )
        self.sub_entry_name = tk.Entry(sub_inputs, width=15)
        self.sub_entry_name.grid(row=0, column=1, padx=2)

        tk.Label(sub_inputs, text="Machine:", bg="#eef2f7").grid(
            row=0, column=2, sticky="w"
        )
        m_box = tk.Frame(sub_inputs, bg="#eef2f7")
        m_box.grid(row=0, column=3, padx=2)
        self.sub_combo_machine = ttk.Combobox(m_box, width=10)
        self.sub_combo_machine.pack(side="left")
        tk.Button(
            m_box,
            text="⚙️",
            font=("Arial", 7),
            command=lambda: self.open_list_manager("Machines", self.data_mgr.machines_db),
        ).pack(side="left", padx=1)

        tk.Label(sub_inputs, text="Operator:", bg="#eef2f7").grid(
            row=0, column=4, sticky="w"
        )
        o_box = tk.Frame(sub_inputs, bg="#eef2f7")
        o_box.grid(row=0, column=5, padx=2)
        self.sub_combo_operator = ttk.Combobox(o_box, width=9)
        self.sub_combo_operator.pack(side="left")
        tk.Button(
            o_box,
            text="👤",
            font=("Arial", 7),
            command=lambda: self.open_list_manager("Operators", self.data_mgr.operators_db),
        ).pack(side="left", padx=1)

        tk.Label(sub_inputs, text="Qty:", bg="#eef2f7").grid(
            row=0, column=6, sticky="w"
        )
        self.sub_entry_qty = tk.Spinbox(sub_inputs, from_=1, to=100000, width=6)
        self.sub_entry_qty.grid(row=0, column=7, padx=2)

        tk.Label(sub_inputs, text="Priority:", bg="#eef2f7").grid(
            row=1, column=0, sticky="w", pady=4
        )
        self.sub_combo_priority = ttk.Combobox(
            sub_inputs, values=self.priorities_db, state="readonly", width=12
        )
        self.sub_combo_priority.grid(row=1, column=1, padx=2, pady=4)
        self.sub_combo_priority.set("Normal")

        tk.Label(sub_inputs, text="Due Date:", bg="#eef2f7").grid(
            row=1, column=2, sticky="w"
        )
        self.sub_entry_date = DateEntry(
            sub_inputs, width=11, date_pattern="dd-mm-yyyy", locale="en_US"
        )
        self.sub_entry_date.grid(row=1, column=3, padx=2)

        tk.Label(sub_inputs, text="Status:", bg="#eef2f7").grid(
            row=1, column=4, sticky="w"
        )
        s_box = tk.Frame(sub_inputs, bg="#eef2f7")
        s_box.grid(row=1, column=5, padx=2)
        self.sub_combo_status = ttk.Combobox(s_box, width=9)
        self.sub_combo_status.pack(side="left")
        tk.Button(
            s_box,
            text="🔄",
            font=("Arial", 7),
            command=lambda: self.open_list_manager("Statuses", self.data_mgr.status_db),
        ).pack(side="left", padx=1)

        tk.Label(sub_inputs, text="MCT:", bg="#eef2f7").grid(
            row=1, column=6, sticky="w"
        )
        self.sub_entry_mct = tk.Entry(sub_inputs, width=10)
        self.sub_entry_mct.grid(row=1, column=7, padx=2)

        tk.Label(sub_inputs, text="File:", bg="#eef2f7").grid(
            row=1, column=8, sticky="w"
        )
        sub_file_box = tk.Frame(sub_inputs, bg="#eef2f7")
        sub_file_box.grid(row=1, column=9, sticky="w")
        self.sub_entry_file = tk.Entry(sub_file_box, width=13)
        self.sub_entry_file.pack(side="left")
        tk.Button(
            sub_file_box, text="...", font=("Arial", 7), command=self.browse_sub_file
        ).pack(side="left", padx=1)

        self.btn_save_sub = tk.Button(
            sub_inputs,
            text="+ Add Sub-part",
            bg="#2980b9",
            fg="white",
            font=("Arial", 8, "bold"),
            command=self.add_or_update_sub_item,
        )
        self.btn_save_sub.grid(row=1, column=10, padx=5)

        # Enable Enter key to move focus to the next input field while adding a new job/sub-part
        try:
            self.combo_client.bind("<Return>", lambda e: self.entry_job.focus_set())
            self.entry_job.bind("<Return>", lambda e: self.entry_po.focus_set())
            self.entry_po.bind("<Return>", lambda e: self.sub_entry_name.focus_set())

            self.sub_entry_name.bind("<Return>", lambda e: self.sub_combo_machine.focus_set())
            self.sub_combo_machine.bind("<Return>", lambda e: self.sub_combo_operator.focus_set())
            self.sub_combo_operator.bind("<Return>", lambda e: self.sub_entry_qty.focus_set())
            self.sub_entry_qty.bind("<Return>", lambda e: self.sub_combo_priority.focus_set())
            self.sub_combo_priority.bind("<Return>", lambda e: self.sub_entry_date.focus_set())
            self.sub_entry_date.bind("<Return>", lambda e: self.sub_combo_status.focus_set())
            self.sub_combo_status.bind("<Return>", lambda e: self.sub_entry_mct.focus_set())
            self.sub_entry_mct.bind("<Return>", lambda e: self.sub_entry_file.focus_set())
            self.sub_entry_file.bind("<Return>", lambda e: self.btn_save_sub.focus_set())
            self.btn_save_sub.bind("<Return>", lambda e: self.add_or_update_sub_item())
        except Exception:
            # Fail silently if any widget isn't present for some reason
            pass

        self.sub_tree = ttk.Treeview(
            sub_frame,
            columns=(
                "sub_no",
                "name",
                "machine",
                "operator",
                "qty",
                "priority",
                "date",
                "status",
                "mct",
                "file",
            ),
            show="headings",
            height=3,
        )
        for col, txt, w in [
            ("sub_no", "Sub-part No.", 90),
            ("name", "Part Name", 120),
            ("machine", "Machine", 110),
            ("operator", "Operator", 90),
            ("qty", "Qty", 50),
            ("priority", "Priority", 70),
            ("date", "Due Date 📅", 85),
            ("status", "Status", 90),
            ("mct", "MCT", 80),
            ("file", "File Name", 130),
        ]:
            self.sub_tree.heading(col, text=txt)
            self.sub_tree.column(col, width=w, anchor="center")
        self.sub_tree.pack(side="left", fill="x", expand=True, pady=3)

        sub_side_btns = tk.Frame(sub_frame, bg="#eef2f7")
        sub_side_btns.pack(side="right", fill="y", padx=3)

        btn_edit_sub = tk.Button(
            sub_side_btns,
            text="✏️\nEdit",
            bg="#d35400",
            fg="white",
            font=("Arial", 8, "bold"),
            command=self.start_edit_sub_item,
        )
        btn_edit_sub.pack(fill="x", pady=1)

        btn_del_sub = tk.Button(
            sub_side_btns,
            text="🗑️\nDel",
            bg="#c0392b",
            fg="white",
            font=("Arial", 8, "bold"),
            command=self.remove_sub_item_from_list,
        )
        btn_del_sub.pack(fill="x", pady=1)

        btn_box = tk.Frame(self.form_frame, bg="#f4f4f9")
        btn_box.pack(fill="x", pady=(5, 0))

        self.btn_add = tk.Button(
            btn_box,
            text="Save Job & All Sub-parts",
            bg="#27ae60",
            fg="white",
            font=("Arial", 9, "bold"),
            command=self.save_job_from_form,
        )
        self.btn_add.pack(side="left", fill="x", expand=True, padx=(0, 2))

        self.btn_cancel = tk.Button(
            btn_box,
            text="Cancel / Close",
            bg="#7f8c8d",
            fg="white",
            font=("Arial", 9, "bold"),
            command=self.hide_form,
        )
        self.btn_cancel.pack(side="left", fill="x", expand=True, padx=(2, 0))

        # --- TABELA GŁÓWNA ---
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview", font=("Arial", 9), rowheight=26)
        style.configure(
            "Treeview.Heading", font=("Arial", 9, "bold"), background="#cbd5e1"
        )

        table_frame = tk.Frame(self.root, bg="#f4f4f9")
        table_frame.pack(fill="both", expand=True, padx=15, pady=5)

        self.columns = (
            "client",
            "date",
            "job",
            "po",
            "sub_no",
            "name",
            "machine",
            "operator",
            "qty",
            "priority",
            "status",
            "mct",
            "file",
        )
        self.tree = ttk.Treeview(
            table_frame, columns=self.columns, show="tree headings", height=14
        )

        self.tree.heading("#0", text=" Order Hierarchy ", anchor="w", command=lambda: self.sort_tree_by_column("#0"))
        self.tree.column("#0", width=180, anchor="w")

        headers = {
            "client": "Client Name 📇",
            "date": "Due Date 📅",
            "job": "Job No.",
            "po": "PO Number",
            "sub_no": "Item No.",
            "name": "Part Name",
            "machine": "Machine ⚙️",
            "operator": "Operator 👤",
            "qty": "Qty",
            "priority": "Priority",
            "status": "Status 🔄",
            "mct": "MCT 📜",
            "file": "Drawing / File",
        }
        widths = {
            "client": 140,
            "date": 90,
            "job": 90,
            "po": 90,
            "sub_no": 90,
            "name": 120,
            "machine": 110,
            "operator": 100,
            "qty": 50,
            "priority": 70,
            "status": 110,
            "mct": 80,
            "file": 140,
        }

        for col in self.columns:
            self.tree.heading(col, text=headers[col], command=lambda c=col: self.sort_tree_by_column(c))
            anchor = (
                "center"
                if col in ("job", "po", "sub_no", "qty", "priority", "date", "status", "mct")
                else "w"
            )
            self.tree.column(col, width=widths[col], anchor=anchor)

        # PODWÓJNE KLIKNIĘCIE WIERSZA (INTELIGENTNA OBSŁUGA)
        self.tree.bind("<Double-1>", self.on_double_click_row)

        self.tree.tag_configure("job_header", font=("Arial", 9, "bold"))

        # Special tags (status tags configured dynamically later)
        self.tree.tag_configure("overdue", background="#f8d7da", foreground="#721c24")
        self.tree.tag_configure("separator", background="#e2e8f0", foreground="#e2e8f0")

        scrollbar = ttk.Scrollbar(
            table_frame, orient="vertical", command=self.tree.yview
        )
        self.tree.configure(yscroll=scrollbar.set)
        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # --- PANEL AKCJI ---
        action_frame = tk.Frame(self.root, bg="#f4f4f9", pady=5)
        action_frame.pack(fill="x", padx=15)

        btn_edit = tk.Button(
            action_frame,
            text="✏️ Edit Full Job",
            bg="#2980b9",
            fg="white",
            font=("Arial", 9, "bold"),
            command=self.start_edit_job,
        )
        btn_edit.pack(side="left", padx=3)

        btn_duplicate = tk.Button(
            action_frame,
            text="📋 Duplicate Job",
            bg="#d35400",
            fg="white",
            font=("Arial", 9, "bold"),
            command=self.duplicate_selected_job,
        )
        btn_duplicate.pack(side="left", padx=3)

        btn_quick_assign = tk.Button(
            action_frame,
            text="⚙️ / 👤 / ⭐ Quick Edit",
            bg="#8e44ad",
            fg="white",
            font=("Arial", 9, "bold"),
            command=self.open_quick_assign_dialog,
        )
        btn_quick_assign.pack(side="left", padx=3)

        btn_change_status = tk.Button(
            action_frame,
            text="🔄 Change Status",
            bg="#16a085",
            fg="white",
            font=("Arial", 9, "bold"),
            command=self.open_status_change_dialog,
        )
        btn_change_status.pack(side="left", padx=3)

        btn_open_file = tk.Button(
            action_frame,
            text="👁️ Open Drawing/File",
            bg="#d35400",
            fg="white",
            font=("Arial", 9, "bold"),
            command=self.open_selected_file,
        )
        btn_open_file.pack(side="left", padx=3)

        btn_open_mct = tk.Button(
            action_frame,
            text="📜 Open MCT Cert",
            bg="#2c3e50",
            fg="white",
            font=("Arial", 9, "bold"),
            command=self.open_selected_mct,
        )
        btn_open_mct.pack(side="left", padx=3)

        btn_delete = tk.Button(
            action_frame,
            text="🗑️ Delete Selected",
            bg="#c0392b",
            fg="white",
            font=("Arial", 9, "bold"),
            command=self.delete_selected,
        )
        btn_delete.pack(side="right", padx=3)

        # --- FOOTER STATS ---
        self.stats_frame = tk.Frame(self.root, bg="#34495e", pady=6)
        self.stats_frame.pack(fill="x", side="bottom")

        self.lbl_stats = tk.Label(
            self.stats_frame,
            text="",
            font=("Arial", 9, "bold"),
            fg="#ecf0f1",
            bg="#34495e",
        )
        self.lbl_stats.pack()

        self.update_clock()
        self.load_data()
        self.check_file_updates()
        self.check_due_date_reminders()

    def check_file_updates(self):
        try:
            if os.path.exists(self.data_mgr.db_file_path):
                mtime = os.path.getmtime(self.data_mgr.db_file_path)
                if (
                    self.data_mgr.last_db_mtime != 0
                    and mtime > self.data_mgr.last_db_mtime
                    and self.editing_job_id is None
                ):
                    self.load_data()
        except Exception:
            pass
        self.root.after(5000, self.check_file_updates)

    def check_due_date_reminders(self):
        try:
            reminded_dates = getattr(self, "_reminded_dates", None)
            if reminded_dates is None:
                reminded_dates = {}
                self._reminded_dates = reminded_dates

            today = datetime.now().date()
            reminder_window = today + timedelta(days=2)
            due_soon_items = []
            overdue_items = []
            for job_no, job_data in self.data_mgr.all_jobs_data.items():
                for item in job_data.get("sub_items", []):
                    try:
                        due_date = datetime.strptime(item.get("date", ""), "%d-%m-%Y").date()
                    except Exception:
                        continue
                    if item.get("status", "") == "Finished":
                        continue
                    if due_date < today:
                        overdue_items.append((job_no, job_data, item))
                    elif due_date <= reminder_window:
                        due_soon_items.append((job_no, job_data, item))

            if overdue_items or due_soon_items:
                self.show_due_date_reminder_window(overdue_items, due_soon_items)
                for job_no, job_data, item in due_soon_items + overdue_items:
                    reminder_key = f"{job_no}:{item['sub_no']}"
                    reminded_dates[reminder_key] = True
        except Exception:
            pass
        self.root.after(60000, self.check_due_date_reminders)

    def open_reminder_window(self):
        overdue_items = []
        due_soon_items = []
        today = datetime.now().date()
        reminder_window = today + timedelta(days=2)
        for job_no, job_data in self.data_mgr.all_jobs_data.items():
            for item in job_data.get("sub_items", []):
                try:
                    due_date = datetime.strptime(item.get("date", ""), "%d-%m-%Y").date()
                except Exception:
                    continue
                if item.get("status", "") == "Finished":
                    continue
                if due_date < today:
                    overdue_items.append((job_no, job_data, item))
                elif due_date <= reminder_window:
                    due_soon_items.append((job_no, job_data, item))

        self.show_due_date_reminder_window(overdue_items, due_soon_items)

    def show_due_date_reminder_window(self, overdue_items, due_soon_items):
        items = []
        for job_no, job_data, item in overdue_items:
            items.append(("Overdue", job_no, job_data, item))
        for job_no, job_data, item in due_soon_items:
            items.append(("Due soon", job_no, job_data, item))

        if not items:
            return

        reminder_window = tk.Toplevel(self.root)
        reminder_window.title("Schedule Reminders")
        reminder_window.geometry("980x420")
        reminder_window.minsize(860, 380)
        reminder_window.transient(self.root)
        reminder_window.grab_set()
        reminder_window.configure(bg="#f8f9fa")

        tk.Label(
            reminder_window,
            text="Schedule Reminder",
            font=("Arial", 14, "bold"),
            bg="#f8f9fa",
        ).pack(pady=(12, 6))

        tk.Label(
            reminder_window,
            text="Below are overdue and due-soon sub-parts. Review the schedule and update any items as needed.",
            font=("Arial", 10),
            bg="#f8f9fa",
        ).pack(pady=(0, 10))

        tree_frame = tk.Frame(reminder_window, bg="#f8f9fa")
        tree_frame.pack(fill="both", expand=True, padx=12, pady=(0, 10))

        columns = ("section", "client", "po", "job", "subpart", "date")
        tree = ttk.Treeview(
            tree_frame,
            columns=columns,
            show="headings",
            selectmode="browse",
        )
        tree.heading("section", text="Section")
        tree.heading("client", text="Client")
        tree.heading("po", text="PO")
        tree.heading("job", text="Job")
        tree.heading("subpart", text="Sub-part")
        tree.heading("date", text="Due Date")
        tree.column("section", width=110, anchor="center")
        tree.column("client", width=240, anchor="w")
        tree.column("po", width=150, anchor="w")
        tree.column("job", width=120, anchor="center")
        tree.column("subpart", width=120, anchor="center")
        tree.column("date", width=120, anchor="center")
        tree.tag_configure("Overdue", background="#f8d7da")
        tree.tag_configure("Due soon", background="#fff3cd")

        for section, job_no, job_data, item in items:
            tree.insert(
                "",
                tk.END,
                values=(
                    section,
                    job_data.get("client", ""),
                    job_data.get("po", ""),
                    job_no,
                    item.get("sub_no", ""),
                    item.get("date", ""),
                ),
                tags=(section,),
            )

        tree.pack(fill="both", expand=True, side="left")
        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        scrollbar.pack(side="right", fill="y")
        tree.configure(yscrollcommand=scrollbar.set)

        footer_frame = tk.Frame(reminder_window, bg="#f8f9fa")
        footer_frame.pack(fill="x", padx=12, pady=(0, 12))

        tk.Label(
            footer_frame,
            text=f"Overdue: {len(overdue_items)}   |   Due soon: {len(due_soon_items)}",
            font=("Arial", 10, "bold"),
            bg="#f8f9fa",
        ).pack(side="left")

        tk.Button(
            footer_frame,
            text="Close",
            bg="#2980b9",
            fg="white",
            font=("Arial", 10, "bold"),
            width=12,
            command=reminder_window.destroy,
        ).pack(side="right")

    def on_double_click_row(self, event=None):
        if event:
            column = self.tree.identify_column(event.x)
            item_id = self.tree.identify_row(event.y)

            if not item_id:
                return

            if column == "#12":  # MCT Column
                self.open_selected_mct()
                return
            elif column == "#13":  # File Column
                self.open_selected_file()
                return
            elif column == "#11":  # Status Column
                self.open_status_change_dialog()
                return
            elif column in ("#5", "#6"):  # Item No / Part Name Column
                self.start_edit_job()
                return

        self.open_quick_assign_dialog()

    def sort_tree_by_column(self, col):
        if self.sort_column == col:
            self.sort_reverse = not self.sort_reverse
        else:
            self.sort_column = col
            self.sort_reverse = False
        self.refresh_tree()

    def open_app_settings_menu(self):
        settings_dialog = tk.Toplevel(self.root)
        settings_dialog.title("App Settings")
        settings_dialog.geometry("280x200")
        settings_dialog.resizable(False, False)
        settings_dialog.grab_set()

        tk.Label(
            settings_dialog,
            text="Application Settings",
            font=("Arial", 11, "bold"),
        ).pack(pady=12)

        tk.Button(
            settings_dialog,
            text="⚙️ Manage Lists",
            width=24,
            command=lambda: [
                settings_dialog.destroy(),
                self.open_lists_manager_menu(),
            ],
        ).pack(pady=4)

        tk.Button(
            settings_dialog,
            text="📁 MCT Folder Location",
            width=24,
            command=lambda: [
                settings_dialog.destroy(),
                self.select_mct_directory(),
            ],
        ).pack(pady=4)

        tk.Button(
            settings_dialog,
            text="💾 Database Location",
            width=24,
            command=lambda: [
                settings_dialog.destroy(),
                self.select_db_file(),
            ],
        ).pack(pady=4)

    def select_db_file(self):
        file_path = filedialog.asksaveasfilename(
            title="Wybierz lub utwórz plik bazy danych JSON",
            defaultextension=".json",
            filetypes=[("Pliki JSON", "*.json"), ("Wszystkie pliki", "*.*")],
            initialdir=os.path.dirname(self.data_mgr.db_file_path),
            initialfile=os.path.basename(self.data_mgr.db_file_path),
        )
        if file_path:
            self.data_mgr.db_file_path = file_path
            save_db_path(self.data_mgr.db_file_path)
            if os.path.exists(self.data_mgr.db_file_path):
                self.load_data()
            else:
                self.save_data()
            messagebox.showinfo(
                "Baza danych", f"Lokalizacja bazy danych zmieniona na:\n{self.data_mgr.db_file_path}"
            )

    def select_mct_directory(self):
        folder = filedialog.askdirectory(
            title="Select MCT Certificates Folder",
            initialdir=self.data_mgr.mct_dir if self.data_mgr.mct_dir else APP_DIR,
        )
        if folder:
            self.data_mgr.mct_dir = folder
            self.save_data()
            messagebox.showinfo(
                "MCT Directory Selected", f"MCT Directory set to:\n{self.data_mgr.mct_dir}"
            )

    def open_selected_mct(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo(
                "Information", "Please select a Sub-part to view its MCT Certificate."
            )
            return

        item_id = sel[0]

        if "_" not in item_id:
            messagebox.showwarning(
                "Selection Error", "Please select a specific Sub-part row (not the Order header)."
            )
            return

        job_no, idx_str = item_id.split("_")
        try:
            idx = int(idx_str)
            sub_item = self.data_mgr.all_jobs_data[job_no]["sub_items"][idx]
            mct_val = sub_item.get("mct", "").strip()
        except (KeyError, IndexError, ValueError):
            messagebox.showerror("Error", "Could not retrieve item data.")
            return

        if not mct_val or mct_val in ("—", "N/A"):
            messagebox.showwarning(
                "No MCT Number", "The selected item does not have an MCT number!"
            )
            return

        if not self.data_mgr.mct_dir or not os.path.exists(self.data_mgr.mct_dir):
            if messagebox.askyesno(
                "MCT Directory Not Set",
                "MCT directory is not set or valid. Would you like to select the folder now?",
            ):
                self.select_mct_directory()
                if not self.data_mgr.mct_dir or not os.path.exists(self.data_mgr.mct_dir):
                    return
            else:
                return

        found_file = None
        mct_clean = os.path.splitext(mct_val)[0].strip().lower()

        exact_path = os.path.join(self.data_mgr.mct_dir, mct_val)
        if os.path.exists(exact_path) and os.path.isfile(exact_path):
            found_file = exact_path
        else:
            extensions = [
                ".pdf",
                ".PDF",
                ".jpg",
                ".jpeg",
                ".png",
                ".tif",
                ".tiff",
                ".doc",
                ".docx",
                ".txt",
            ]
            for ext in extensions:
                test_path = os.path.join(self.data_mgr.mct_dir, mct_val + ext)
                if os.path.exists(test_path):
                    found_file = test_path
                    break

        if not found_file:
            try:
                all_files = os.listdir(self.data_mgr.mct_dir)
                for f in all_files:
                    if mct_clean in f.lower():
                        found_file = os.path.join(self.data_mgr.mct_dir, f)
                        break
            except Exception as e:
                messagebox.showerror("Error", f"Error scanning MCT directory:\n{str(e)}")
                return

        if found_file and os.path.exists(found_file):
            try:
                if sys.platform.startswith("win"):
                    os.startfile(found_file)
                elif sys.platform.startswith("darwin"):
                    subprocess.run(["open", found_file])
                else:
                    subprocess.run(["xdg-open", found_file])
            except Exception as e:
                messagebox.showerror(
                    "Error", f"Failed to open MCT certificate file:\n{str(e)}"
                )
        else:
            messagebox.showwarning(
                "Certificate Not Found",
                f"Could not find a certificate matching '{mct_val}' in folder:\n{self.data_mgr.mct_dir}",
            )

    def open_lists_manager_menu(self):
        menu_dialog = tk.Toplevel(self.root)
        menu_dialog.title("Manage Lists")
        menu_dialog.geometry("260x230")
        menu_dialog.resizable(False, False)
        menu_dialog.grab_set()

        tk.Label(
            menu_dialog,
            text="Select List to Manage",
            font=("Arial", 10, "bold"),
        ).pack(pady=10)

        tk.Button(
            menu_dialog,
            text="📇 Clients List",
            width=22,
            command=lambda: [
                menu_dialog.destroy(),
                self.open_list_manager("Clients", self.data_mgr.clients_db),
            ],
        ).pack(pady=4)

        tk.Button(
            menu_dialog,
            text="⚙️ Machines List",
            width=22,
            command=lambda: [
                menu_dialog.destroy(),
                self.open_list_manager("Machines", self.data_mgr.machines_db),
            ],
        ).pack(pady=4)

        tk.Button(
            menu_dialog,
            text="👤 Operators List",
            width=22,
            command=lambda: [
                menu_dialog.destroy(),
                self.open_list_manager("Operators", self.data_mgr.operators_db),
            ],
        ).pack(pady=4)

        tk.Button(
            menu_dialog,
            text="🔄 Statuses List",
            width=22,
            command=lambda: [
                menu_dialog.destroy(),
                self.open_list_manager("Statuses", self.data_mgr.status_db),
            ],
        ).pack(pady=4)

    def open_list_manager(self, title_name, item_set):
        if title_name.lower() == "statuses":
            ListManagerDialog(
                self.root,
                title_name,
                item_set,
                self.update_all_comboboxes,
                color_map=self.data_mgr.status_colors,
            )
        else:
            ListManagerDialog(
                self.root, title_name, item_set, self.update_all_comboboxes
            )

    def update_all_comboboxes(self):
        """Aktualizuje listy wyboru oraz natychmiast zapisuje dane do JSON."""
        clients_sorted = sorted(list(self.data_mgr.clients_db))
        machines_sorted = sorted(list(self.data_mgr.machines_db))
        operators_sorted = sorted(list(self.data_mgr.operators_db))
        statuses_sorted = sorted(list(self.data_mgr.status_db))

        self.combo_client["values"] = clients_sorted
        self.sub_combo_machine["values"] = machines_sorted
        self.sub_combo_operator["values"] = operators_sorted
        self.sub_combo_status["values"] = statuses_sorted

        self.combo_filter_machine["values"] = ["All"] + machines_sorted
        self.combo_filter_status["values"] = ["All"] + statuses_sorted
        self.combo_filter_client["values"] = ["All"] + clients_sorted
        # Reconfigure status tags/colors from persisted map and save
        try:
            self.configure_status_tags()
        except Exception:
            pass
        self.save_data()

    def configure_status_tags(self):
        """Configure Treeview tags for statuses using colors from DataManager."""
        # Start with any existing mapping from data_mgr
        merged = dict(self.data_mgr.status_colors or {})

        # Ensure every status exists in mapping (default to white background if missing)
        for s in self.data_mgr.status_db:
            if s not in merged:
                merged[s] = "#ffffff"

        # Save to instance and configure tags
        self.status_color_map = {}
        for s, col in merged.items():
            s_norm = str(s).strip().lower()
            tag_name = f"status_{s_norm.replace(' ', '_')}"
            bg = col or "#ffffff"
            fg = "#000000"
            try:
                self.tree.tag_configure(tag_name, background=bg, foreground=fg)
            except Exception:
                pass
            self.status_color_map[s_norm] = (bg, fg)

    def update_clock(self):
        now = datetime.now().strftime("%d-%m-%Y %H:%M:%S")
        self.clock_label.config(text=now)
        self.root.after(1000, self.update_clock)

    def load_data(self):
        try:
            self.data_mgr.load_data()
        except RuntimeError as e:
            messagebox.showerror("Error", str(e))
        self.update_all_comboboxes()
        self.refresh_tree()

    def save_data(self):
        """Trwały zapis do pliku bazy JSON za pomocą DataManager."""
        try:
            self.data_mgr.save_data()
        except RuntimeError as e:
            messagebox.showerror("Error", str(e))

    def open_add_form(self):
        self.clear_form()
        self.editing_job_id = None
        self.btn_toggle_form.config(state="disabled")
        self.form_frame.pack(fill="x", padx=15, pady=5, before=self.tree.master)
        self._reminded_dates = getattr(self, "_reminded_dates", {})

    def hide_form(self):
        self.form_frame.pack_forget()
        self.btn_toggle_form.config(state="normal")
        self.clear_form()

    def clear_form(self):
        self.combo_client.set("")
        self.entry_job.config(state="normal")
        self.entry_job.delete(0, tk.END)
        self.entry_po.delete(0, tk.END)
        self.temp_sub_items = []
        self.editing_sub_idx = None
        self.clear_sub_inputs()
        self.refresh_sub_tree()

    def clear_sub_inputs(self):
        self.sub_entry_name.delete(0, tk.END)
        self.sub_combo_machine.set("")
        self.sub_combo_operator.set("")
        self.sub_entry_qty.delete(0, tk.END)
        self.sub_entry_qty.insert(0, "1")
        self.sub_combo_priority.set("Normal")
        self.sub_combo_status.set("In Progress")
        self.sub_entry_mct.delete(0, tk.END)
        self.sub_entry_file.delete(0, tk.END)
        self.editing_sub_idx = None
        self.btn_save_sub.config(text="+ Add Sub-part", bg="#2980b9")

    def update_sub_item_numbers(self):
        job_no = self.entry_job.get().strip() or "JOB"
        for idx, item in enumerate(self.temp_sub_items, 1):
            item["sub_no"] = f"{job_no}/{idx}"
        self.refresh_sub_tree()

    def browse_sub_file(self):
        f_path = filedialog.askopenfilename(
            title="Select Drawing or Document",
            filetypes=[("All Files", "*.*"), ("PDF files", "*.pdf"), ("CAD files", "*.dwg;*.dxf")],
        )
        if f_path:
            self.sub_entry_file.delete(0, tk.END)
            self.sub_entry_file.insert(0, f_path)

    def add_or_update_sub_item(self):
        job_no = self.entry_job.get().strip() or "JOB"
        name = self.sub_entry_name.get().strip()
        if not name:
            messagebox.showwarning("Warning", "Part Name is required!")
            return

        qty_str = self.sub_entry_qty.get().strip()
        if not qty_str.isdigit() or int(qty_str) <= 0:
            messagebox.showwarning("Warning", "Quantity (Qty) must be a positive integer!")
            return

        m_val = self.sub_combo_machine.get().strip()
        o_val = self.sub_combo_operator.get().strip()
        s_val = self.sub_combo_status.get().strip() or "In Progress"

        if m_val:
            self.data_mgr.machines_db.add(m_val)
        if o_val:
            self.data_mgr.operators_db.add(o_val)
        if s_val:
            self.data_mgr.status_db.add(s_val)

        sub_item_data = {
            "name": name,
            "machine": m_val,
            "operator": o_val,
            "qty": qty_str,
            "priority": self.sub_combo_priority.get(),
            "date": self.sub_entry_date.get_date().strftime("%d-%m-%Y"),
            "status": s_val,
            "mct": self.sub_entry_mct.get().strip(),
            "file": self.sub_entry_file.get().strip(),
        }

        if self.editing_sub_idx is not None:
            sub_item_data["sub_no"] = self.temp_sub_items[self.editing_sub_idx]["sub_no"]
            self.temp_sub_items[self.editing_sub_idx] = sub_item_data
        else:
            sub_item_data["sub_no"] = f"{job_no}/{len(self.temp_sub_items) + 1}"
            self.temp_sub_items.append(sub_item_data)

        self.update_all_comboboxes()
        self.save_data()
        self.clear_sub_inputs()
        self.refresh_sub_tree()

    def refresh_sub_tree(self):
        for item in self.sub_tree.get_children():
            self.sub_tree.delete(item)
        for idx, it in enumerate(self.temp_sub_items):
            self.sub_tree.insert(
                "",
                "end",
                iid=str(idx),
                values=(
                    it["sub_no"],
                    it["name"],
                    it["machine"],
                    it["operator"],
                    it["qty"],
                    it["priority"],
                    it["date"],
                    it["status"],
                    it.get("mct", ""),
                    os.path.basename(it["file"]) if it["file"] else "",
                ),
            )

    def start_edit_sub_item(self):
        sel = self.sub_tree.selection()
        if not sel:
            return
        idx = int(sel[0])
        it = self.temp_sub_items[idx]
        self.editing_sub_idx = idx

        self.sub_entry_name.delete(0, tk.END)
        self.sub_entry_name.insert(0, it["name"])
        self.sub_combo_machine.set(it["machine"])
        self.sub_combo_operator.set(it["operator"])
        self.sub_entry_qty.delete(0, tk.END)
        self.sub_entry_qty.insert(0, it["qty"])
        self.sub_combo_priority.set(it["priority"])

        try:
            d_obj = datetime.strptime(it["date"], "%d-%m-%Y")
            self.sub_entry_date.set_date(d_obj)
        except Exception:
            pass

        self.sub_combo_status.set(it["status"])
        self.sub_entry_mct.delete(0, tk.END)
        self.sub_entry_mct.insert(0, it.get("mct", ""))
        self.sub_entry_file.delete(0, tk.END)
        self.sub_entry_file.insert(0, it["file"])

        self.btn_save_sub.config(text="Update Sub-part", bg="#d35400")

    def remove_sub_item_from_list(self):
        sel = self.sub_tree.selection()
        if not sel:
            return
        if not messagebox.askyesno("Confirm Delete", "Are you sure you want to remove this sub-part?"):
            return
        idx = int(sel[0])
        del self.temp_sub_items[idx]
        self.update_sub_item_numbers()
        self.clear_sub_inputs()

    def save_job_from_form(self):
        client = self.combo_client.get().strip()
        job_no = self.entry_job.get().strip()
        po = self.entry_po.get().strip()

        if not client or not job_no:
            messagebox.showwarning(
                "Warning", "Client Name and Job No. are required!"
            )
            return

        if not self.temp_sub_items:
            messagebox.showwarning(
                "Warning", "Please add at least one Sub-part!"
            )
            return

        if self.editing_job_id is None and job_no in self.data_mgr.all_jobs_data:
            messagebox.showerror(
                "Error", f"Job Number '{job_no}' already exists!"
            )
            return

        self.data_mgr.clients_db.add(client)
        for it in self.temp_sub_items:
            if it["machine"]:
                self.data_mgr.machines_db.add(it["machine"])
            if it["operator"]:
                self.data_mgr.operators_db.add(it["operator"])
            if it["status"]:
                self.data_mgr.status_db.add(it["status"])

        self.data_mgr.all_jobs_data[job_no] = {
            "client": client,
            "po": po,
            "sub_items": self.temp_sub_items,
        }

        self.update_all_comboboxes()
        self.save_data()
        self.hide_form()
        self.refresh_tree()

    def start_edit_job(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("Information", "Please select a Job or Sub-part to edit.")
            return

        item_id = sel[0]
        if "_" in item_id:
            job_no = item_id.split("_")[0]
        else:
            job_no = item_id

        if job_no not in self.data_mgr.all_jobs_data:
            return

        job_data = self.data_mgr.all_jobs_data[job_no]
        self.open_add_form()
        self.editing_job_id = job_no

        self.combo_client.set(job_data["client"])
        self.entry_job.insert(0, job_no)
        self.entry_job.config(state="disabled")
        self.entry_po.insert(0, job_data["po"])

        self.temp_sub_items = [dict(it) for it in job_data["sub_items"]]
        self.refresh_sub_tree()

    def duplicate_selected_job(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("Information", "Please select a Job or Sub-part to duplicate.")
            return

        item_id = sel[0]
        job_no = item_id.split("_")[0] if "_" in item_id else item_id

        if job_no not in self.data_mgr.all_jobs_data:
            return

        orig_data = self.data_mgr.all_jobs_data[job_no]
        new_job_no = f"{job_no}_COPY"

        self.open_add_form()
        self.combo_client.set(orig_data["client"])
        self.entry_job.insert(0, new_job_no)
        self.entry_po.insert(0, orig_data["po"])

        copied_subs = []
        for idx, it in enumerate(orig_data["sub_items"], 1):
            c_it = dict(it)
            c_it["sub_no"] = f"{new_job_no}/{idx}"
            copied_subs.append(c_it)

        self.temp_sub_items = copied_subs
        self.refresh_sub_tree()

    def delete_selected(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("Information", "Please select an item to delete.")
            return

        if not messagebox.askyesno(
            "Confirm Delete", "Are you sure you want to delete selected item(s)?"
        ):
            return

        for item_id in sel:
            if "_" in item_id:
                job_no, idx_str = item_id.split("_")
                idx = int(idx_str)
                if job_no in self.data_mgr.all_jobs_data:
                    sub_list = self.data_mgr.all_jobs_data[job_no]["sub_items"]
                    if 0 <= idx < len(sub_list):
                        del sub_list[idx]
                        if not sub_list:
                            del self.data_mgr.all_jobs_data[job_no]
                        else:
                            for i, it in enumerate(sub_list, 1):
                                it["sub_no"] = f"{job_no}/{i}"
            else:
                if item_id in self.data_mgr.all_jobs_data:
                    del self.data_mgr.all_jobs_data[item_id]

        self.save_data()
        self.refresh_tree()

    def refresh_tree(self):
        for item in self.tree.get_children():
            self.tree.delete(item)

        search_q = self.search_var.get().strip().lower()
        f_machine = self.filter_machine_var.get()
        f_status = self.filter_status_var.get()
        f_client = self.filter_client_var.get()
        f_date = self.filter_date_var.get().strip()

        total_sub_items = 0
        status_counts = {}
        today_str = datetime.now().strftime("%d-%m-%Y")

        sorted_jobs = list(self.data_mgr.all_jobs_data.items())

        # Sortowanie według wybranej kolumny jeśli włączone
        if self.sort_column:
            col = self.sort_column
            def get_sort_key(item):
                job_no, job_data = item
                if col == "date":
                    subs = job_data.get("sub_items", [])
                    if subs:
                        # Sortuj po najwcześniejszej dacie z podzadań
                        dates = []
                        for sub in subs:
                            try:
                                dates.append(datetime.strptime(sub.get("date", ""), "%d-%m-%Y").date())
                            except Exception:
                                continue
                        if dates:
                            return min(dates)
                    return datetime.max.date()
                if col in ("client", "po", "job", "#0"):
                    val = job_data.get(col, job_no)
                else:
                    subs = job_data.get("sub_items", [])
                    val = subs[0].get(col, "") if subs else ""
                return str(val).lower()
            sorted_jobs.sort(key=get_sort_key, reverse=self.sort_reverse)

        for job_no, job_data in sorted_jobs:
            client = job_data["client"]
            po = job_data["po"]
            sub_items = job_data["sub_items"]

            matching_subs = []
            for idx, it in enumerate(sub_items):
                if f_machine != "All" and it["machine"] != f_machine:
                    continue
                if f_status != "All" and it["status"] != f_status:
                    continue
                if f_client != "All" and client != f_client:
                    continue
                if f_date and it["date"] != f_date:
                    continue

                if search_q:
                    haystack = f"{client} {job_no} {po} {it['sub_no']} {it['name']} {it['machine']} {it['operator']} {it['status']} {it.get('mct','')}".lower()
                    if search_q not in haystack:
                        continue

                matching_subs.append((idx, it))

            if matching_subs:
                has_high_priority = any(
                    str(it.get("priority", "")).strip().lower() == "high"
                    for _, it in matching_subs
                )
                parent_text = f"⭐ 📦 Order: {job_no}" if has_high_priority else f"📦 Order: {job_no}"

                parent_node = self.tree.insert(
                    "",
                    "end",
                    iid=job_no,
                    text=parent_text,
                    values=(
                        client,
                        "",
                        job_no,
                        po,
                        f"({len(matching_subs)} sub-parts)",
                        "—",
                        "—",
                        "—",
                        "—",
                        "—",
                        "—",
                        "—",
                        "—",
                    ),
                    tags=("job_header",),
                    open=True,
                )

                for idx, it in matching_subs:
                    child_id = f"{job_no}_{idx}"
                    st = it["status"]
                    tags = []
                    st_norm = str(st).strip().lower()

                    # Add tag for known status colors
                    status_tag = f"status_{st_norm.replace(' ', '_')}"
                    if st_norm in getattr(self, 'status_color_map', {}):
                        tags.append(status_tag)

                    # Mark overdue items (unless finished)
                    try:
                        if it["date"] < today_str and st_norm != "finished":
                            tags.append("overdue")
                    except Exception:
                        pass

                    self.tree.insert(
                        parent_node,
                        "end",
                        iid=child_id,
                        text=f"  📄 {it['sub_no']}",
                        values=(
                            "",
                            it["date"],
                            "",
                            "",
                            it["sub_no"],
                            it["name"],
                            it["machine"],
                            it["operator"],
                            it["qty"],
                            it["priority"],
                            it["status"],
                            it.get("mct", ""),
                            os.path.basename(it["file"]) if it["file"] else "",
                        ),
                        tags=tuple(tags),
                    )

                    total_sub_items += 1
                    status_counts[st] = status_counts.get(st, 0) + 1

                self.tree.insert(
                    parent_node,
                    "end",
                    values=("", "", "", "", "", "", "", "", "", "", "", "", ""),
                    tags=("separator",),
                )

        stats_str = f"Total Active Sub-parts: {total_sub_items}  |  " + "  |  ".join(
            [f"{k}: {v}" for k, v in status_counts.items()]
        )
        self.lbl_stats.config(text=stats_str)

    def apply_filters(self):
        self.refresh_tree()

    def reset_filters(self):
        self.search_var.set("")
        self.filter_machine_var.set("All")
        self.filter_status_var.set("All")
        self.filter_client_var.set("All")
        self.filter_date_var.set("")
        self.refresh_tree()

    def open_selected_file(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("Information", "Please select a Sub-part file to open.")
            return

        item_id = sel[0]
        if "_" not in item_id:
            messagebox.showinfo("Information", "Please select a specific Sub-part row.")
            return

        job_no, idx_str = item_id.split("_")
        idx = int(idx_str)
        f_path = self.data_mgr.all_jobs_data[job_no]["sub_items"][idx]["file"]

        if f_path and os.path.exists(f_path):
            try:
                if sys.platform.startswith("win"):
                    os.startfile(f_path)
                elif sys.platform.startswith("darwin"):
                    subprocess.run(["open", f_path])
                else:
                    subprocess.run(["xdg-open", f_path])
            except Exception as e:
                messagebox.showerror("Error", f"Failed to open file:\n{str(e)}")
        else:
            messagebox.showwarning(
                "Warning", "File path is empty or file does not exist!"
            )

    def open_quick_assign_dialog(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("Information", "Please select an Order or Sub-part row first.")
            return

        item_id = sel[0]
        is_full_job = "_" not in item_id

        if is_full_job:
            job_no = item_id
            sub_items = self.data_mgr.all_jobs_data[job_no]["sub_items"]
            title_txt = f"Quick Edit ALL Sub-parts in {job_no}"
            init_m = sub_items[0]["machine"] if sub_items else ""
            init_o = sub_items[0]["operator"] if sub_items else ""
            init_p = sub_items[0]["priority"] if sub_items else "Normal"
            init_mct = sub_items[0].get("mct", "") if sub_items else ""
        else:
            job_no, idx_str = item_id.split("_")
            idx = int(idx_str)
            sub_item = self.data_mgr.all_jobs_data[job_no]["sub_items"][idx]
            title_txt = f"Quick Edit: {sub_item['sub_no']}"
            init_m = sub_item["machine"]
            init_o = sub_item["operator"]
            init_p = sub_item["priority"]
            init_mct = sub_item.get("mct", "")

        dlg = tk.Toplevel(self.root)
        dlg.title("Quick Edit Machine / Operator / Priority / MCT")
        dlg.geometry("360x290")
        dlg.resizable(False, False)
        dlg.grab_set()

        tk.Label(
            dlg, text=title_txt, font=("Arial", 10, "bold")
        ).pack(pady=8)

        f_inputs = tk.Frame(dlg)
        f_inputs.pack(pady=5)

        tk.Label(f_inputs, text="Machine:").grid(row=0, column=0, sticky="w", pady=4)
        c_m = ttk.Combobox(
            f_inputs, values=sorted(list(self.data_mgr.machines_db)), width=16
        )
        c_m.grid(row=0, column=1, padx=5, pady=4)
        c_m.set(init_m)

        tk.Label(f_inputs, text="Operator:").grid(row=1, column=0, sticky="w", pady=4)
        c_o = ttk.Combobox(
            f_inputs, values=sorted(list(self.data_mgr.operators_db)), width=16
        )
        c_o.grid(row=1, column=1, padx=5, pady=4)
        c_o.set(init_o)

        tk.Label(f_inputs, text="Priority:").grid(row=2, column=0, sticky="w", pady=4)
        c_p = ttk.Combobox(
            f_inputs, values=self.priorities_db, state="readonly", width=16
        )
        c_p.grid(row=2, column=1, padx=5, pady=4)
        c_p.set(init_p)

        tk.Label(f_inputs, text="MCT:").grid(row=3, column=0, sticky="w", pady=4)
        c_mct = tk.Entry(f_inputs, width=18)
        c_mct.grid(row=3, column=1, padx=5, pady=4)
        c_mct.insert(0, init_mct)

        def save_quick():
            m_val = c_m.get().strip()
            o_val = c_o.get().strip()
            p_val = c_p.get().strip()
            mct_val = c_mct.get().strip()
            if m_val:
                self.data_mgr.machines_db.add(m_val)
            if o_val:
                self.data_mgr.operators_db.add(o_val)

            if is_full_job:
                for sub in self.data_mgr.all_jobs_data[job_no]["sub_items"]:
                    sub["machine"] = m_val
                    sub["operator"] = o_val
                    sub["priority"] = p_val
                    sub["mct"] = mct_val
            else:
                sub_item["machine"] = m_val
                sub_item["operator"] = o_val
                sub_item["priority"] = p_val
                sub_item["mct"] = mct_val

            self.update_all_comboboxes()
            self.refresh_tree()
            dlg.destroy()

        tk.Button(
            dlg,
            text="Save Assignment",
            bg="#27ae60",
            fg="white",
            font=("Arial", 9, "bold"),
            command=save_quick,
        ).pack(pady=10)

    def open_status_change_dialog(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("Information", "Please select a Sub-part row first.")
            return

        selected_sub_items = []
        for item_id in sel:
            if "_" in item_id:
                job_no, idx_str = item_id.split("_")
                idx = int(idx_str)
                selected_sub_items.append(self.data_mgr.all_jobs_data[job_no]["sub_items"][idx])

        if not selected_sub_items:
            messagebox.showinfo("Information", "Please select specific Sub-part row(s).")
            return

        dlg = tk.Toplevel(self.root)
        dlg.title("Change Status")
        dlg.geometry("300x140")
        dlg.resizable(False, False)
        dlg.grab_set()

        lbl_text = f"Change Status for {len(selected_sub_items)} selected Sub-part(s)"
        tk.Label(
            dlg, text=lbl_text, font=("Arial", 10, "bold")
        ).pack(pady=10)

        c_s = ttk.Combobox(dlg, values=sorted(list(self.data_mgr.status_db)), width=18)
        c_s.pack(pady=5)
        c_s.set(selected_sub_items[0]["status"])

        def save_st():
            s_val = c_s.get().strip()
            if s_val:
                self.data_mgr.status_db.add(s_val)
                for sub in selected_sub_items:
                    sub["status"] = s_val
                self.update_all_comboboxes()
                self.refresh_tree()
            dlg.destroy()

        tk.Button(
            dlg,
            text="Update Status",
            bg="#16a085",
            fg="white",
            font=("Arial", 9, "bold"),
            command=save_st,
        ).pack(pady=10)

    def export_to_excel(self):
        if Workbook is None:
            messagebox.showerror(
                "Export Error",
                "Excel export is unavailable because openpyxl is not installed.",
            )
            return

        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Export Schedule Data to Excel",
        )
        if not filename:
            return

        headers = [
            "Client",
            "Due Date",
            "Job No",
            "PO Number",
            "Sub-part No",
            "Part Name",
            "Machine",
            "Operator",
            "Qty",
            "Priority",
            "Status",
            "MCT",
            "File Path",
        ]

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Schedule"
            ws.append(headers)

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="4B0082")
            header_alignment = Alignment(horizontal="center", vertical="center")

            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            status_fill_map = {
                st.lower(): color.replace("#", "")
                for st, color in self.data_mgr.status_colors.items()
                if color
            }
            row_alignment = Alignment(vertical="center", wrap_text=True)

            for job_num, job_data in self.data_mgr.all_jobs_data.items():
                for it in job_data["sub_items"]:
                    due_date = None
                    try:
                        due_date = datetime.strptime(it["date"], "%d-%m-%Y").date()
                    except Exception:
                        due_date = it["date"]

                    ws.append(
                        [
                            job_data["client"],
                            due_date,
                            job_num,
                            job_data["po"],
                            it["sub_no"],
                            it["name"],
                            it["machine"],
                            it["operator"],
                            it["qty"],
                            it["priority"],
                            it["status"],
                            it.get("mct", ""),
                            it["file"],
                        ]
                    )
                    current_row = ws.max_row
                    status_key = str(it.get("status", "")).strip().lower()
                    fill_color = status_fill_map.get(status_key)
                    if fill_color:
                        row_fill = PatternFill(fill_type="solid", fgColor=fill_color)
                        for col_idx in range(1, len(headers) + 1):
                            cell = ws.cell(row=current_row, column=col_idx)
                            cell.fill = row_fill

            ws.freeze_panes = ws["A2"]
            ws.auto_filter.ref = ws.dimensions

            # Auto-fit columns and apply row alignment
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    if cell.row == 1:
                        cell.alignment = header_alignment
                    else:
                        cell.alignment = row_alignment

                    value = cell.value
                    if value is None:
                        continue
                    cell_length = len(str(value))
                    if cell_length > max_length:
                        max_length = cell_length
                adjusted_width = min(max_length + 3, 60)
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(filename)
            messagebox.showinfo(
                "Export Successful", f"Data exported successfully to:\n{filename}"
            )
        except Exception as e:
            messagebox.showerror(
                "Export Error", f"Failed to export data:\n{str(e)}"
            )

    def show_column_mapping_dialog(self, headers, targets):
        """Show a dialog allowing the user to map worksheet headers to target fields.

        Returns a dict mapping each target -> selected header name (or empty string).
        Returns None if the user cancels.
        """
        dlg = tk.Toplevel(self.root)
        dlg.title("Map Excel Columns")
        dlg.geometry("680x420")
        dlg.transient(self.root)
        dlg.grab_set()

        tk.Label(
            dlg,
            text="Map columns from the Excel file to application fields:",
            font=("Arial", 10, "bold"),
        ).pack(pady=(8, 6), padx=8, anchor="w")

        frame = tk.Frame(dlg)
        frame.pack(fill="both", expand=True, padx=8, pady=4)

        left = tk.Frame(frame)
        left.pack(side="left", fill="both", expand=True)
        right = tk.Frame(frame, width=240)
        right.pack(side="right", fill="y", padx=(8,0))

        # Show detected headers on the right for clarity
        tk.Label(right, text="Detected headers:", font=("Arial", 9, "bold")).pack(anchor="nw", pady=(2,4))
        lb_frame = tk.Frame(right)
        lb_frame.pack(fill="both", expand=True)
        header_listbox = tk.Listbox(lb_frame, height=18, activestyle="none")
        header_scroll = ttk.Scrollbar(lb_frame, orient="vertical", command=header_listbox.yview)
        header_listbox.configure(yscrollcommand=header_scroll.set)
        header_listbox.pack(side="left", fill="both", expand=True)
        header_scroll.pack(side="right", fill="y")
        for h in headers:
            header_listbox.insert(tk.END, h)

        entries = {}
        for i, key in enumerate(targets):
            lbl = tk.Label(left, text=key.title() + ":", anchor="w")
            lbl.grid(row=i, column=0, sticky="w", padx=4, pady=3)
            cb = ttk.Combobox(left, values=headers, state="readonly", width=60)
            cb.grid(row=i, column=1, sticky="w", padx=4, pady=3)
            # try to preselect a reasonable match
            sel = ""
            for h in headers:
                if key in h.lower():
                    sel = h
                    break
            if sel:
                cb.set(sel)
            entries[key] = cb

        btn_frame = tk.Frame(dlg)
        btn_frame.pack(fill="x", pady=(6, 10), padx=8)

        result = {}

        def on_ok():
            for k, comb in entries.items():
                v = comb.get().strip()
                result[k] = v if v else ""
            dlg.destroy()

        def on_cancel():
            result.clear()
            dlg.destroy()

        tk.Button(btn_frame, text="OK", width=10, command=on_ok).pack(side="right", padx=6)
        tk.Button(btn_frame, text="Cancel", width=10, command=on_cancel).pack(side="right")

        dlg.wait_window()
        if not result:
            return None
        return result

    def import_from_excel(self):
        if load_workbook is None:
            messagebox.showerror(
                "Import Error",
                "Excel import is unavailable because openpyxl is not installed.",
            )
            return

        filename = filedialog.askopenfilename(
            defaultextension=".xlsx",
            filetypes=[
                ("Excel files", "*.xlsx;*.xlsm;*.xltx;*.xltm"),
                ("All files", "*.*"),
            ],
            title="Import Schedule Data from Excel",
        )
        if not filename:
            return

        try:
            wb = load_workbook(filename, data_only=True)
            ws = wb.active
            # Try to detect the header row among the first 10 rows (take the row with most non-empty cells)
            best_row = None
            best_count = -1
            for r in range(1, min(11, ws.max_row + 1)):
                row_cells = ws[r]
                row_vals = [str(cell.value).strip() if cell.value is not None else "" for cell in row_cells]
                non_empty = sum(1 for v in row_vals if v)
                if non_empty > best_count:
                    best_count = non_empty
                    best_row = row_vals
            if not best_row or best_count == 0:
                messagebox.showerror(
                    "Import Error",
                    "Could not detect header row in Excel file. Make sure headers are in the first rows.",
                )
                return
            header_row = best_row

            # Fields we will ask the user to map
            targets = [
                "client",
                "due date",
                "job no",
                "po number",
                "sub-part no",
                "part name",
                "machine",
                "operator",
                "qty",
                "priority",
                "status",
                "mct",
                "file path",
            ]

            mapping = self.show_column_mapping_dialog(header_row, targets)
            if mapping is None:
                return

            # build header_map: target -> index
            header_map = {}
            for key, header_name in mapping.items():
                if header_name:
                    try:
                        idx = header_row.index(header_name)
                        header_map[key] = idx
                    except ValueError:
                        # header not found, ignore
                        pass

            if "job no" not in header_map:
                messagebox.showerror(
                    "Import Error",
                    "You must map the 'Job No' column to import data.",
                )
                return

            imported_rows = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue

                def val(name):
                    idx = header_map.get(name)
                    return row[idx] if idx is not None and idx < len(row) else ""

                client = str(val("client")).strip()
                job_no = str(val("job no")).strip()
                po = str(val("po number")).strip()
                sub_no = str(val("sub-part no")).strip()
                part_name = str(val("part name")).strip()
                machine = str(val("machine")).strip()
                operator = str(val("operator")).strip()
                qty = val("qty")
                priority = str(val("priority")).strip()
                status = str(val("status")).strip()
                mct = str(val("mct")).strip()
                file_path = str(val("file path")).strip()

                if not job_no:
                    continue

                due_date_val = val("due date")
                due_date = ""
                if isinstance(due_date_val, (datetime,)):
                    due_date = due_date_val.strftime("%d-%m-%Y")
                elif due_date_val is not None:
                    due_date = str(due_date_val).strip()

                if job_no not in self.data_mgr.all_jobs_data:
                    self.data_mgr.all_jobs_data[job_no] = {
                        "client": client,
                        "po": po,
                        "sub_items": [],
                    }
                else:
                    existing_job = self.data_mgr.all_jobs_data[job_no]
                    if client:
                        existing_job["client"] = client
                    if po:
                        existing_job["po"] = po

                sub_item = {
                    "sub_no": sub_no or f"{job_no}/1",
                    "name": part_name,
                    "machine": machine,
                    "operator": operator,
                    "qty": qty if qty is not None else "",
                    "priority": priority,
                    "date": due_date,
                    "status": status,
                    "mct": mct,
                    "file": file_path,
                }

                job_data = self.data_mgr.all_jobs_data[job_no]
                existing_subs = {it["sub_no"]: it for it in job_data.get("sub_items", [])}
                if sub_item["sub_no"] in existing_subs:
                    existing_subs[sub_item["sub_no"]].update(sub_item)
                else:
                    job_data.setdefault("sub_items", []).append(sub_item)

                if client:
                    self.data_mgr.clients_db.add(client)
                if machine:
                    self.data_mgr.machines_db.add(machine)
                if operator:
                    self.data_mgr.operators_db.add(operator)
                if status:
                    self.data_mgr.status_db.add(status)
                    if status not in self.data_mgr.status_colors:
                        self.data_mgr.status_colors[status] = "#ffffff"

                imported_rows += 1

            self.update_all_comboboxes()
            self.save_data()
            self.refresh_tree()
            messagebox.showinfo(
                "Import Successful",
                f"Imported {imported_rows} rows from Excel file."
            )
        except Exception as e:
            messagebox.showerror(
                "Import Error",
                f"Failed to import Excel data:\n{str(e)}",
            )


if __name__ == "__main__":
    root = tk.Tk()
    root.state("zoomed")

    # Pokazujemy planszę informacyjną przez 3000 ms (3 sekundy)
    splash = SplashScreen(root, delay_ms=3000)

    # Inicjalizacja aplikacji
    app = ScheduleApp(root)

    root.mainloop()